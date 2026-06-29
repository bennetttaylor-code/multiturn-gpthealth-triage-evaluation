#!/usr/bin/env python3
"""
GPT Triage Analysis – Statistical Analysis Script
==================================================
Compares GPT-assisted triage decisions (Natural and Multiturn conditions)
and Clinician-Adjudicated against Nurse Triage and Clinician-Adjudicated as primary reference standards.

Usage:
    python analysis.py --input "Analysis Copy Here.xlsx" --output ./output/

Steps:
    0  – Descriptives & data cleaning
    2  – Primary LLM analysis (Natural & Multiturn vs. Nurse Triage)
    3  – Clinician-Adjudicated benchmark vs. Nurse Triage
    4  – Sensitivity analysis (conservative / aggressive resolution)
    5  – Subgroup analysis by dataset
    6  – Natural vs. Multiturn paired comparison
    6b – LLM vs. Clinician-Adjudicated paired comparison (Bonferroni-corrected)
    7  – Safety-critical error analysis
    8  – Prompt count as predictor
    9  – Chief complaint category analysis
    10 – Grader-level reliability
    11 – Direction-of-error asymmetry (binomial test)
    12 – Master summary table + Key Findings
"""

import argparse
import os
import sys
import warnings
import numpy as np
import pandas as pd
from scipy.stats import wilcoxon, mannwhitneyu, spearmanr, norm, chi2, binomtest
from sklearn.metrics import cohen_kappa_score
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# =====================================================================
# CLI
# =====================================================================
parser = argparse.ArgumentParser(description='GPT Triage Statistical Analysis')
parser.add_argument('--input',  required=True,
                    help='Path to input Excel file (sheet named C)')
parser.add_argument('--output', default='./output',
                    help='Output directory (default: ./output)')
args = parser.parse_args()

INPUT_PATH  = args.input
OUTPUT_DIR  = args.output
OUTPUT_FILE = os.path.join(OUTPUT_DIR, 'triage_analysis_results.xlsx')
os.makedirs(OUTPUT_DIR, exist_ok=True)

# =====================================================================
# CONSTANTS
# =====================================================================
ORD_MAP = {'A': 0, 'B': 1, 'C': 2, 'D': 3}
ORD_INV = {0: 'A', 1: 'B', 2: 'C', 3: 'D'}
LETTERS = ['A', 'B', 'C', 'D']

COMPLAINT_CAT = {
    'Abdomen Pain': 'Gastrointestinal', 'Abdominal Pain - Male': 'Gastrointestinal',
    'Abdominal Pain - female': 'Gastrointestinal', 'abdominal pain ': 'Gastrointestinal',
    'Nausea': 'Gastrointestinal', 'nausea': 'Gastrointestinal',
    'Vomiting': 'Gastrointestinal', 'vomiting': 'Gastrointestinal',
    'Diarrhea and trouble urinating': 'Gastrointestinal', 'diarrhea': 'Gastrointestinal',
    'Rectal bleeding': 'Gastrointestinal', 'rectal symptoms': 'Gastrointestinal',
    'Constipation': 'Gastrointestinal', 'constipation ': 'Gastrointestinal',
    'Stools - Unusual Color': 'Gastrointestinal', 'Jaundice': 'Gastrointestinal',
    'Mild-Moderate abdominal pain (per triage but notes actually say severe pain), constant for > 2 hours ': 'Gastrointestinal',
    'Chest Pain': 'Cardiovascular', 'Chest or Rib Pain': 'Cardiovascular',
    'Palpitations': 'Cardiovascular', 'Heart Rate and Heartbeat Questions': 'Cardiovascular',
    'Blood Pressure - High': 'Cardiovascular', 'High blood pressure': 'Cardiovascular',
    'HIGH BLOOD PRESSURE': 'Cardiovascular', 'high blood pressure': 'Cardiovascular',
    'Leg Swelling and Edema': 'Cardiovascular', 'leg swelling': 'Cardiovascular',
    'leg swellling': 'Cardiovascular',
    'Thigh, calf or ankle swelling and only one side': 'Cardiovascular',
    'Ankle Swelling': 'Cardiovascular',
    'Calf or leg pain, one sided, present 1 hour': 'Cardiovascular',
    'Cough': 'Respiratory', 'cough': 'Respiratory',
    'Shortness of Breath': 'Respiratory', 'Asthma Attack': 'Respiratory',
    'difficulty breathing': 'Respiratory',
    'Mild difficulty breathing, new onset, worse than normal ': 'Respiratory',
    'Nose Stuffiness or Congestion': 'Respiratory',
    'Headache': 'Neurological/Neuropsychiatric', 'headache': 'Neurological/Neuropsychiatric',
    'Dizziness': 'Neurological/Neuropsychiatric', 'dizziness': 'Neurological/Neuropsychiatric',
    'Loss of Consciousness': 'Neurological/Neuropsychiatric',
    'Neurologic Deficit': 'Neurological/Neuropsychiatric',
    'Neurologic deficit': 'Neurological/Neuropsychiatric',
    'Neurological Deficit': 'Neurological/Neuropsychiatric',
    'Muscle Jerks or Twitches': 'Neurological/Neuropsychiatric',
    'Depression': 'Neurological/Neuropsychiatric',
    'Suicide Concerns': 'Neurological/Neuropsychiatric',
    'head injury': 'Neurological/Neuropsychiatric',
    'Arm Pain': 'Musculoskeletal', 'Back Pain': 'Musculoskeletal',
    'Back pain': 'Musculoskeletal', 'back pain': 'Musculoskeletal',
    'back pain ': 'Musculoskeletal', 'Hip Pain': 'Musculoskeletal',
    'Knee Pain': 'Musculoskeletal', 'Knee Swelling': 'Musculoskeletal',
    'Leg Pain': 'Musculoskeletal', 'Foot Pain': 'Musculoskeletal',
    'Foot pain': 'Musculoskeletal', 'Hand Pain': 'Musculoskeletal',
    'hand pain': 'Musculoskeletal', 'left hand pain and swelling': 'Musculoskeletal',
    'Rib Pain': 'Musculoskeletal', 'rib pain ': 'Musculoskeletal',
    'shoulder pain ': 'Musculoskeletal', 'Toe Injury': 'Musculoskeletal',
    'toe pain ': 'Musculoskeletal', 'right toe pain ': 'Musculoskeletal',
    'groin pain': 'Musculoskeletal', 'neck pain ': 'Musculoskeletal',
    'muscle spasms': 'Musculoskeletal', 'Fall': 'Musculoskeletal', 'fall ': 'Musculoskeletal',
    'Severe ankle pain not improved 2 hours after medication ': 'Musculoskeletal',
    'Severe arm pain and not better after medications and ice ': 'Musculoskeletal',
    'Severe back pain, unable to do normal activities, not improved 2 hours after pain meds': 'Musculoskeletal',
    'Redness or Rash': 'Dermatological', 'Rash-widespread': 'Dermatological',
    'rash': 'Dermatological', 'Skin Blisters': 'Dermatological',
    'Skin Lesion - Moles or Growths': 'Dermatological',
    'Localized purple or blood colored spots, not from injury or friction and no fever': 'Dermatological',
    'Purple or blood colored rash, no fever, user sounds good to triager, drug rash suspected, started new medication ': 'Dermatological',
    'Rash - Purple Spots or Dots': 'Dermatological', 'Cold sore': 'Dermatological',
    'Face Swelling': 'Dermatological', 'Neck Swelling': 'Dermatological',
    'skin growth ': 'Dermatological', 'skin lump': 'Dermatological',
    'itching': 'Dermatological', 'Cut or Laceration': 'Dermatological',
    'Cuts and Lacerations': 'Dermatological', 'sting': 'Dermatological',
    'Eyelid Swelling': 'Dermatological',
    'Hematuria': 'Genitourinary', 'Urine - Blood In': 'Genitourinary',
    'Urine-blood in': 'Genitourinary', 'blood in urine': 'Genitourinary',
    'Urinary symptoms': 'Genitourinary', 'Urination Pain, Female': 'Genitourinary',
    'urination pain ': 'Genitourinary',
    'Fever': 'Infectious/Systemic', 'Fever >100 AND bedridden': 'Infectious/Systemic',
    'Fever >101 AND >60 years old': 'Infectious/Systemic', 'Fever >103': 'Infectious/Systemic',
    'Fever of 101F and age 60 ': 'Infectious/Systemic', 'fever': 'Infectious/Systemic',
    'High risk patient (>64 years old, diabetes, heart disease, weak immune system) AND flu exposure within 7 days AND cold symptoms': 'Infectious/Systemic',
    'High risk patient (>64 years old, diabetes, heart disease, weak immune system) AND symptoms are worsening': 'Infectious/Systemic',
    'Fatigue': 'Infectious/Systemic', 'Generalized Weakness': 'Infectious/Systemic',
    'Weakness (Generalized) and Fatigue': 'Infectious/Systemic', 'weakness': 'Infectious/Systemic',
    'Anaphylaxis': 'Infectious/Systemic',
    'Blood sugar problem or diabetes': 'Infectious/Systemic',
    'Diabetes - High Blood Sugar': 'Infectious/Systemic',
    'blood sugar problem or diabetes': 'Infectious/Systemic',
    'diabetes': 'Infectious/Systemic', 'high blood sugar': 'Infectious/Systemic',
    'Vaginal Discharge': 'OB/GYN', 'Vaginal Pain or Irritation': 'OB/GYN',
    'Vaginal bleeding': 'OB/GYN', 'vaginal bleeding': 'OB/GYN',
    'Ear Pain': 'Other/Unclassified', 'Eye - Redness': 'Other/Unclassified',
    'Eye Pain': 'Other/Unclassified', 'eye pain': 'Other/Unclassified',
    'Sore Throat': 'Other/Unclassified', 'Throat Pain': 'Other/Unclassified',
    'sore throat': 'Other/Unclassified', 'Tooth or Gum Pain': 'Other/Unclassified',
    'Nosebleed': 'Other/Unclassified', 'earwax problems': 'Other/Unclassified',
    'Mouth Pain': 'Other/Unclassified', 'Mouth Symptoms': 'Other/Unclassified',
    'mouth sores': 'Other/Unclassified', 'Medication question': 'Other/Unclassified',
    'medication refill ': 'Other/Unclassified', 'inhaler refill': 'Other/Unclassified',
    'psych referral ': 'Other/Unclassified', 'severe pain after surgery': 'Other/Unclassified',
    'post op wound': 'Other/Unclassified', 'wound care': 'Other/Unclassified',
    'wound check': 'Other/Unclassified',
}

# =====================================================================
# HELPER FUNCTIONS
# =====================================================================

def to_range(val):
    """Parse value to (lo, hi) ordinal bounds. Single A/B/C/D → lo=hi. Range → lo<hi."""
    if pd.isna(val):
        return (np.nan, np.nan)
    v = str(val).strip()
    if v in ORD_MAP:
        o = ORD_MAP[v]
        return (o, o)
    for sep in ['/', '-']:
        if sep in v:
            parts = [p.strip() for p in v.split(sep)]
            if all(p in ORD_MAP for p in parts):
                ords = [ORD_MAP[p] for p in parts]
                return (min(ords), max(ords))
    return (np.nan, np.nan)

def weighted_kappa_ci(y1, y2, n_boot=1000, seed=42):
    """Cohen's weighted kappa (linear) with bootstrapped 95% CI."""
    mask = (~pd.isna(y1)) & (~pd.isna(y2))
    a = np.array(y1)[mask].astype(float)
    b = np.array(y2)[mask].astype(float)
    if len(a) < 4:
        return np.nan, np.nan, np.nan
    try:
        k = cohen_kappa_score(a.astype(int), b.astype(int), weights='linear', labels=[0,1,2,3])
    except Exception:
        return np.nan, np.nan, np.nan
    rng = np.random.RandomState(seed)
    boots = []
    n = len(a)
    for _ in range(n_boot):
        idx = rng.randint(0, n, n)
        try:
            boots.append(cohen_kappa_score(a[idx].astype(int), b[idx].astype(int),
                                           weights='linear', labels=[0,1,2,3]))
        except Exception:
            pass
    if len(boots) < 50:
        return k, np.nan, np.nan
    return k, float(np.percentile(boots, 2.5)), float(np.percentile(boots, 97.5))

def wilson_ci(n_agree, n_total, alpha=0.05):
    # Exact (Clopper-Pearson) binomial 95% CI for proportions, to match the
    # confidence intervals reported in the manuscript. Returns (p, lo, hi).
    if n_total == 0:
        return np.nan, np.nan, np.nan
    p = n_agree / n_total
    ci = binomtest(int(n_agree), int(n_total), 0.5).proportion_ci(
        confidence_level=1 - alpha, method='exact')
    return p, ci.low, ci.high

def spearman_ci(rho, n, alpha=0.05):
    if n <= 3:
        return np.nan, np.nan
    z = np.arctanh(rho)
    se = 1 / np.sqrt(n - 3)
    zc = norm.ppf(1 - alpha / 2)
    return np.tanh(z - zc * se), np.tanh(z + zc * se)

def compute_metrics(lo_s, hi_s, gold_s):
    """
    Compute all metrics using range-based minimum-distance scoring.
    lo_s / hi_s: lower/upper bounds of decision range (lo=hi for unambiguous values).
    gold_s: single ordinal reference value (0-3). NaN rows are dropped.
    """
    lo   = np.array(lo_s,   dtype=float)
    hi   = np.array(hi_s,   dtype=float)
    gold = np.array(gold_s, dtype=float)
    mask = ~(np.isnan(lo) | np.isnan(hi) | np.isnan(gold))
    lo, hi, gold = lo[mask], hi[mask], gold[mask]
    n = len(lo)
    if n == 0:
        return {'n': 0}

    in_range = (lo <= gold) & (gold <= hi)
    below    = hi < gold   # range entirely below gold → under-triage
    above    = lo > gold   # range entirely above gold → over-triage
    absdist  = np.where(in_range, 0.0, np.where(below, gold - hi, lo - gold))
    signed   = np.where(in_range, 0.0, np.where(below, -(gold - hi), lo - gold))

    n_agree = int(in_range.sum())
    agree_p, agree_lo, agree_hi = wilson_ci(n_agree, n)
    dist_counts = {i: int((absdist == i).sum()) for i in range(4)}
    dist_pcts   = {i: dist_counts[i] / n * 100  for i in range(4)}
    mean_dist   = float(absdist.mean())
    sd_dist     = float(absdist.std(ddof=1)) if n > 1 else 0.0

    n_under = int((signed < 0).sum())
    n_over  = int((signed > 0).sum())
    under_by_step = {i: int((signed == -i).sum()) for i in range(1, 4)}
    over_by_step  = {i: int((signed ==  i).sum()) for i in range(1, 4)}

    # Oracle assignment: clamp gold to [lo, hi] — most favorable per-row kappa assignment
    oracle = np.clip(gold, lo, hi)
    kappa, klo, khi = weighted_kappa_ci(oracle, gold)

    return {
        'n': n, 'n_agree': n_agree,
        'agree_pct': agree_p * 100, 'agree_ci_lo': agree_lo * 100, 'agree_ci_hi': agree_hi * 100,
        'dist_counts': dist_counts, 'dist_pcts': dist_pcts,
        'mean_dist': mean_dist, 'sd_dist': sd_dist,
        'n_under': n_under, 'under_pct': n_under / n * 100,
        'under_by_step': under_by_step,
        'n_over': n_over, 'over_pct': n_over / n * 100,
        'over_by_step': over_by_step,
        'kappa': kappa, 'kappa_lo': klo, 'kappa_hi': khi,
    }

def metrics_to_rows(m):
    if m.get('n', 0) == 0:
        return [('N', 0)]
    rows = [
        ('N cases', m['n']),
        ('Exact Agreement (%)',
         f"{m['agree_pct']:.1f}% (95% CI: {m['agree_ci_lo']:.1f}–{m['agree_ci_hi']:.1f}%)"),
        ('Distance 0 steps (%)', f"{m['dist_pcts'][0]:.1f}%"),
        ('Distance 1 step (%)',  f"{m['dist_pcts'][1]:.1f}%"),
        ('Distance 2 steps (%)', f"{m['dist_pcts'][2]:.1f}%"),
        ('Distance 3 steps (%)', f"{m['dist_pcts'][3]:.1f}%"),
        ('Mean Clinical Distance (SD)', f"{m['mean_dist']:.3f} ({m['sd_dist']:.3f})"),
        ('Under-triage rate – any (%)', f"{m['under_pct']:.1f}%"),
        ('  Under-triage 1 step (%)',  f"{m['under_by_step'][1]/m['n']*100:.1f}%"),
        ('  Under-triage 2 steps (%)', f"{m['under_by_step'][2]/m['n']*100:.1f}%"),
        ('  Under-triage 3 steps (%)', f"{m['under_by_step'][3]/m['n']*100:.1f}%"),
        ('Over-triage rate – any (%)', f"{m['over_pct']:.1f}%"),
        ('  Over-triage 1 step (%)',  f"{m['over_by_step'][1]/m['n']*100:.1f}%"),
        ('  Over-triage 2 steps (%)', f"{m['over_by_step'][2]/m['n']*100:.1f}%"),
        ('  Over-triage 3 steps (%)', f"{m['over_by_step'][3]/m['n']*100:.1f}%"),
    ]
    kap_str = f"{m['kappa']:.3f}" if not np.isnan(m.get('kappa', np.nan)) else 'N/A'
    ci_str  = (f"({m['kappa_lo']:.3f}, {m['kappa_hi']:.3f})"
               if not np.isnan(m.get('kappa_lo', np.nan)) else '')
    rows.append(("Cohen's Weighted Kappa (95% CI)", f"{kap_str} {ci_str}"))
    return rows

def make_confusion(lo_s, hi_s, gold_s):
    """4×4 confusion matrix. Oracle (clamp gold→range) assigns cell for range decisions."""
    conf = pd.DataFrame(0, index=LETTERS, columns=LETTERS)
    lo   = np.array(lo_s,   dtype=float)
    hi   = np.array(hi_s,   dtype=float)
    gold = np.array(gold_s, dtype=float)
    for l, h, g in zip(lo, hi, gold):
        if np.isnan(l) or np.isnan(h) or np.isnan(g):
            continue
        oracle = int(np.clip(g, l, h))
        conf.at[ORD_INV[oracle], ORD_INV[int(g)]] += 1
    conf.index.name = 'Decision \\ Reference'
    conf['Total'] = conf.sum(axis=1)
    conf.loc['Total'] = conf.sum(axis=0)
    return conf

def mcnemar_test(agree1, agree2):
    a  = np.array(agree1, dtype=bool)
    b  = np.array(agree2, dtype=bool)
    n01 = int((~a & b).sum())
    n10 = int((a & ~b).sum())
    if n01 + n10 == 0:
        return np.nan, 1.0
    stat = (abs(n01 - n10) - 1)**2 / (n01 + n10)
    return float(stat), float(chi2.sf(stat, df=1))

def cohens_d_paired(x, y):
    diff = np.array(x, dtype=float) - np.array(y, dtype=float)
    return float(diff.mean() / diff.std(ddof=1)) if diff.std(ddof=1) > 0 else np.nan

def kappa_label(k):
    if np.isnan(k): return 'N/A'
    if k < 0:       return 'Poor'
    if k < 0.20:    return 'Slight'
    if k < 0.40:    return 'Fair'
    if k < 0.60:    return 'Moderate'
    if k < 0.80:    return 'Substantial'
    return 'Almost Perfect'

def row_dist(lo_s, hi_s, gold_s):
    lo, hi, g = np.array(lo_s, float), np.array(hi_s, float), np.array(gold_s, float)
    in_r  = (lo <= g) & (g <= hi)
    below = hi < g
    return np.where(in_r, 0.0, np.where(below, g - hi, lo - g))

def row_signed(lo_s, hi_s, gold_s):
    lo, hi, g = np.array(lo_s, float), np.array(hi_s, float), np.array(gold_s, float)
    in_r  = (lo <= g) & (g <= hi)
    below = hi < g
    return np.where(in_r, 0.0, np.where(below, -(g - hi), lo - g))

def parse_col_range(series):
    parsed = series.apply(to_range)
    return parsed.apply(lambda x: x[0]), parsed.apply(lambda x: x[1])

def resolve_pair_single(primary, reviewer, method='conservative'):
    if pd.isna(primary) or pd.isna(reviewer):
        return primary if not pd.isna(primary) else reviewer
    if primary == reviewer:
        return primary
    return max(primary, reviewer) if method == 'conservative' else min(primary, reviewer)

# =====================================================================
# DATA LOADING & CLEANING
# =====================================================================
print(f"Loading data from: {INPUT_PATH}")
df_raw = pd.read_excel(INPUT_PATH, sheet_name=0, header=0)
df_raw.columns = [str(c).strip() for c in df_raw.columns]

COL_MAP = {
    'DATASET': 'dataset', 'Prompt': 'prompt', 'Grader': 'grader',
    'Patient Chief Complaint': 'complaint',
    'Clinician Consensus': 'cc_raw',
    'GPT Natural Prompt Count': 'nat_count',
    'GPT Natural Triage Decision (A-D)': 'nat_triage_raw',
    'GPT Natural Reviewer': 'nat_reviewer_raw',
    'GPT Multiturn Prompt Count': 'mt_count',
    'GPT Multiturn Triage Decision (A-D)': 'mt_triage_raw',
    'GPT Multiturn Reviewer': 'mt_reviewer_raw',
}
df = df_raw.rename(columns=COL_MAP)[list(COL_MAP.values())].copy()
nurse_col = [c for c in df_raw.columns if 'Nurse' in c][0]
df['gold_raw'] = df_raw[nurse_col]

for col in ['cc_raw', 'nat_triage_raw', 'nat_reviewer_raw', 'mt_triage_raw',
            'mt_reviewer_raw', 'gold_raw']:
    df[col] = df[col].astype(str).str.strip().replace('nan', np.nan)

# Fill 1 known missing Natural Reviewer (Dataset 2, Prompt 49)
missing_nat_rev = df[df['nat_reviewer_raw'].isna()].index.tolist()
if missing_nat_rev:
    print(f"  Filling missing Natural Reviewer at row(s) {missing_nat_rev} with 'C'")
    df.loc[missing_nat_rev, 'nat_reviewer_raw'] = 'C'

# Parse all columns to (lo, hi) ranges
df['cc_lo'],        df['cc_hi']        = parse_col_range(df['cc_raw'])
df['nat_lo'],       df['nat_hi']       = parse_col_range(df['nat_triage_raw'])
df['nat_rev_lo'],   df['nat_rev_hi']   = parse_col_range(df['nat_reviewer_raw'])
df['mt_lo'],        df['mt_hi']        = parse_col_range(df['mt_triage_raw'])
df['mt_rev_lo'],    df['mt_rev_hi']    = parse_col_range(df['mt_reviewer_raw'])
df['gold_lo'],      df['gold_hi']      = parse_col_range(df['gold_raw'])

df['nat_ord']     = df['nat_hi']
df['mt_ord']      = df['mt_hi']
df['nat_rev_ord'] = df['nat_rev_hi']
df['mt_rev_ord']  = df['mt_rev_hi']
df['gold_ord']    = df['gold_hi']
df['cc_cons_ord'] = df['cc_hi']
df['cc_agg_ord']  = df['cc_lo']

df['nat_resolved_cons'] = [resolve_pair_single(p, r, 'conservative')
                            for p, r in zip(df['nat_ord'], df['nat_rev_ord'])]
df['nat_resolved_agg']  = [resolve_pair_single(p, r, 'aggressive')
                            for p, r in zip(df['nat_lo'],  df['nat_rev_lo'])]
df['mt_resolved_cons']  = [resolve_pair_single(p, r, 'conservative')
                            for p, r in zip(df['mt_ord'],  df['mt_rev_ord'])]
df['mt_resolved_agg']   = [resolve_pair_single(p, r, 'aggressive')
                            for p, r in zip(df['mt_lo'],   df['mt_rev_lo'])]

df['category'] = df['complaint'].map(COMPLAINT_CAT).fillna('Other/Unclassified')

df['nat_dist']   = row_dist(df['nat_lo'],  df['nat_hi'],  df['gold_ord'])
df['mt_dist']    = row_dist(df['mt_lo'],   df['mt_hi'],   df['gold_ord'])
df['cc_dist']    = row_dist(df['cc_lo'],   df['cc_hi'],   df['gold_ord'])
df['nat_signed'] = row_signed(df['nat_lo'], df['nat_hi'], df['gold_ord'])
df['mt_signed']  = row_signed(df['mt_lo'],  df['mt_hi'],  df['gold_ord'])
df['cc_signed']  = row_signed(df['cc_lo'],  df['cc_hi'],  df['gold_ord'])
df['nat_signed_cc'] = row_signed(df['nat_lo'], df['nat_hi'], df['cc_cons_ord'])
df['mt_signed_cc']  = row_signed(df['mt_lo'],  df['mt_hi'],  df['cc_cons_ord'])

print(f"Data loaded: {len(df)} rows across {df['dataset'].nunique()} datasets, "
      f"{df['grader'].nunique()} graders")

# =====================================================================
# EXCEL WRITER SETUP
# =====================================================================
SHEET_FONT = 'Arial'   # consistent font for all sheets

def style_sheet(ws, title_row=1, header_rows=None):
    # Scientific/booktabs style: no gridlines, no fills, horizontal rules only,
    # consistent serif font, bold title, bold+ruled column headers.
    thin = Side(style='thin', color='000000')
    med  = Side(style='medium', color='000000')
    ws.sheet_view.showGridLines = False
    header_set = set(header_rows or [])

    def as_colheader(r):                  # bold, ruled, left label / centred values
        for cell in r:
            cell.font = Font(name=SHEET_FONT, bold=True, size=11)
            cell.border = Border(top=med, bottom=thin)
        r[0].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        for cell in r[1:]:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def as_data(r):
        for cell in r:
            cell.font = Font(name=SHEET_FONT, size=11)
            cell.alignment = Alignment(horizontal='left', vertical='center')
        for cell in r[1:]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def as_section(r):                    # single-cell section sub-header
        r[0].font = Font(name=SHEET_FONT, bold=True, size=11)
        r[0].alignment = Alignment(horizontal='left', vertical='center')

    prev_section = False
    for row in ws.iter_rows():
        ridx = row[0].row
        n_filled = sum(1 for c in row if c.value not in (None, ''))
        single = (n_filled == 1 and row[0].value not in (None, '') and len(str(row[0].value)) < 110)
        cur_section = False
        if ridx == title_row:
            for cell in row:
                cell.font = Font(name=SHEET_FONT, bold=True, size=13)
                cell.alignment = Alignment(horizontal='left', vertical='center')
        elif ridx in header_set and n_filled >= 2:        # explicit multi-column header
            as_colheader(row)
        elif (ridx in header_set and n_filled < 2) or single:  # section sub-header
            as_data(row)
            as_section(row)
            if ridx in header_set:
                row[0].border = Border(bottom=thin)
            cur_section = True
        elif prev_section and n_filled >= 2:              # column header right after a section title
            as_colheader(row)
        else:
            as_data(row)
        prev_section = cur_section
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 12), 60)

writer = pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl')

# =====================================================================
# TABLE 1: CASE CHARACTERISTICS BY DATASET
# =====================================================================
print("Building Table 1...")

def _t1_pctn(x, n): return f"{int(x)} ({x/n*100:.0f}%)" if n > 0 else '—'
def _t1_med_iqr(s):
    a = np.array(s, float); a = a[~np.isnan(a)]
    if len(a) == 0: return '—'
    return f"{np.median(a):.0f} [{np.percentile(a,25):.0f}–{np.percentile(a,75):.0f}]"

_t1_subs = [df[df['dataset'] == 1], df[df['dataset'] == 2], df[df['dataset'] == 3], df]
_t1_cols = [f'Clinically-Authored Vignettes (n={len(_t1_subs[0])})',
            f'Emergency Department (n={len(_t1_subs[1])})',
            f'Nurse Triage (n={len(_t1_subs[2])})',
            f'Overall (n={len(_t1_subs[3])})']
_t1_rows = []
def _t1_row(label, fn): _t1_rows.append([label] + [fn(s) for s in _t1_subs])
def _t1_hdr(label):     _t1_rows.append([label, '', '', '', ''])
_t1_row('Cases, n',            lambda s: f"{len(s)}")
_t1_hdr('Nurse Triage acuity, n (%)')
for _k in [0, 1, 2, 3]:
    _t1_row(f'    {ORD_INV[_k]}', lambda s, k=_k: _t1_pctn((s['gold_ord'] == k).sum(), len(s)))
_t1_hdr('Clinician-Adjudicated acuity, n (%)')
for _k in [0, 1, 2, 3]:
    _t1_row(f'    {ORD_INV[_k]}', lambda s, k=_k: _t1_pctn((s['cc_cons_ord'] == k).sum(), len(s)))
_t1_hdr('Chief complaint category, n (%)')
for _c in df['category'].value_counts().index.tolist():
    _t1_row(f'    {_c}', lambda s, c=_c: _t1_pctn((s['category'] == c).sum(), len(s)))
_t1_hdr('Prompts to triage, median [IQR]')
_t1_row('    Natural condition',   lambda s: _t1_med_iqr(s['nat_count']))
_t1_row('    Multiturn condition', lambda s: _t1_med_iqr(s['mt_count']))

_ed  = df_raw[df_raw['DATASET'] == 2]
_age = pd.to_numeric(_ed['Age'], errors='coerce').dropna()
_sex = _ed['Gender'].astype(str).str.strip()
_nF, _nM = int((_sex == 'F').sum()), int((_sex == 'M').sum()); _ns = _nF + _nM
_t1_foot = ("Age and sex were recorded only for the Emergency Department subset (n=76): "
            f"age mean {_age.mean():.0f} years (SD {_age.std():.0f}; range {_age.min():.0f}–{_age.max():.0f}); "
            f"sex {_nF} female ({_nF/_ns*100:.0f}%), {_nM} male ({_nM/_ns*100:.0f}%). "
            "Age and sex were not collected for the Clinically-Authored Vignettes or Nurse-Triage datasets.")

sn_t1 = 'Table 1'
rt1 = 0
pd.DataFrame([['TABLE 1. Case Characteristics by Dataset (N = 255)']]).to_excel(
    writer, sheet_name=sn_t1, startrow=rt1, index=False, header=False); rt1 += 2
pd.DataFrame(_t1_rows, columns=['Characteristic'] + _t1_cols).to_excel(
    writer, sheet_name=sn_t1, startrow=rt1, index=False); rt1 += len(_t1_rows) + 2
foot_row = rt1 + 1  # Excel row of the footnote
pd.DataFrame([[_t1_foot]]).to_excel(
    writer, sheet_name=sn_t1, startrow=rt1, index=False, header=False)

# --- Booktabs (scientific journal) styling for the Table 1 tab ---
ws_t1 = writer.sheets[sn_t1]
ws_t1.sheet_view.showGridLines = False
_serif      = SHEET_FONT
_hdr_excel  = 3                       # column-header row
_first_data = 4                       # first data row
_last_data  = 3 + len(_t1_rows)       # last data row
_thick = Side(style='medium', color='000000')
_thin  = Side(style='thin',   color='000000')
# base font + alignment
for row in ws_t1.iter_rows():
    for cell in row:
        cell.font = Font(name=_serif, size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
# title
ws_t1['A1'].font = Font(name=_serif, size=13, bold=True)
# column header row: bold, centred values, top (thick) + bottom (thin) rules
for cell in ws_t1[_hdr_excel]:
    cell.font = Font(name=_serif, size=10, bold=True)
    cell.alignment = Alignment(horizontal=('left' if cell.column == 1 else 'center'),
                               vertical='center', wrap_text=True)
    cell.border = Border(top=_thick, bottom=_thin)
# data rows (each row is [label, v1, v2, v3, v4]; group headers have empty values)
for i, rowvals in enumerate(_t1_rows):
    label = rowvals[0]
    is_h  = all((v == '' or v is None) for v in rowvals[1:])
    er = _first_data + i
    a_cell = ws_t1.cell(row=er, column=1)
    if is_h:                          # group header: bold, no indent
        a_cell.font = Font(name=_serif, size=10, bold=True)
    else:                             # sub-item: indent + centred values
        a_cell.value = label.strip()
        a_cell.alignment = Alignment(horizontal='left', vertical='center',
                                     indent=(1 if label.startswith('    ') else 0))
        for col in range(2, 6):
            ws_t1.cell(row=er, column=col).alignment = Alignment(
                horizontal='center', vertical='center')
# bottom rule under last data row
for col in range(1, 6):
    ws_t1.cell(row=_last_data, column=col).border = Border(bottom=_thick)
# footnote
fcell = ws_t1.cell(row=foot_row, column=1)
fcell.font = Font(name=_serif, size=8, italic=True, color='333333')
fcell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
# column widths
ws_t1.column_dimensions['A'].width = 34
for c in ['B', 'C', 'D', 'E']:
    ws_t1.column_dimensions[c].width = 16
print("  Table 1 complete.")

# =====================================================================
# STEP 0: DESCRIPTIVES
# =====================================================================
print("Running Step 0...")
total = len(df)
s0_rows = []
s0_rows.append(('SECTION 0a: Row Counts', '', ''))
s0_rows.append(('Total rows', total, ''))
for ds in [1, 2, 3]:
    s0_rows.append((f'  Dataset {ds}', int((df['dataset']==ds).sum()), ''))
for g in sorted(df['grader'].unique()):
    s0_rows.append((f'  Grader {g}', int((df['grader']==g).sum()), ''))

decision_cols = {
    'Natural Triage Decision': ('nat_lo', 'nat_hi'),
    'Natural Reviewer Decision': ('nat_rev_lo', 'nat_rev_hi'),
    'Clinician-Adjudicated':      ('cc_lo', 'cc_hi'),
    'Multiturn Triage Decision':('mt_lo', 'mt_hi'),
    'Multiturn Reviewer Decision':('mt_rev_lo', 'mt_rev_hi'),
    'Nurse Triage Decision':    ('gold_lo', 'gold_hi'),
}
s0_rows.append(('', '', ''))
s0_rows.append(('SECTION 0b: Missingness (post-imputation)', '', ''))
s0_rows.append(('Column', 'N Missing', '% Missing'))
for label, (lo_col, _) in decision_cols.items():
    n_miss = int(df[lo_col].isna().sum())
    s0_rows.append((label, n_miss, f"{n_miss/total*100:.1f}%"))

s0_rows.append(('', '', ''))
s0_rows.append(('SECTION 0c: Non-Standard Values', '', ''))
raw_check = {
    'Clinician-Adjudicated (raw)': 'cc_raw',
    'Natural Triage (raw)': 'nat_triage_raw',
    'Natural Reviewer (raw)': 'nat_reviewer_raw',
    'Multiturn Triage (raw)': 'mt_triage_raw',
    'Multiturn Reviewer (raw)': 'mt_reviewer_raw',
    'Nurse Triage (raw)': 'gold_raw',
}
for label, col in raw_check.items():
    non_std = [v for v in df[col].dropna().unique() if str(v).strip() not in LETTERS]
    if non_std:
        for v in non_std:
            cnt = int((df[col] == v).sum())
            lo, hi = to_range(v)
            s0_rows.append((label, f"'{v}' → range [{ORD_INV.get(int(lo),'?')},{ORD_INV.get(int(hi),'?')}]", f"N={cnt} rows"))
    else:
        s0_rows.append((label, 'All values A/B/C/D – standard', ''))
s0_rows.append(('Note', 'Range values scored with min-distance (reference within range = distance 0)', ''))
s0_rows.append(('Note', 'Missing Natural Reviewer (Dataset 2, Prompt 49) filled with C per instruction', ''))

s0_rows.append(('', '', ''))
s0_rows.append(('SECTION 0d: Frequency Distributions', '', ''))
single_cols = {
    'Natural Triage': 'nat_ord', 'Natural Reviewer': 'nat_rev_ord',
    'Clinician-Adjudicated (conservative/hi)': 'cc_cons_ord',
    'Multiturn Triage': 'mt_ord', 'Multiturn Reviewer': 'mt_rev_ord',
    'Nurse Triage (Gold)': 'gold_ord',
}
for label, col in single_cols.items():
    s0_rows.append((label, 'Count', '%'))
    vc = df[col].value_counts().reindex([0,1,2,3], fill_value=0)
    for k, letter in ORD_INV.items():
        s0_rows.append((f'  {letter}', int(vc[k]), f"{vc[k]/total*100:.1f}%"))

s0_rows.append(('', '', ''))
s0_rows.append(('SECTION 0e: Nurse Triage Decisions by Dataset', '', ''))
for ds in [1, 2, 3]:
    n_ds = int((df['dataset']==ds).sum())
    s0_rows.append((f'Dataset {ds} (N={n_ds})', 'Count', '%'))
    vc = df[df['dataset']==ds]['gold_ord'].value_counts().reindex([0,1,2,3], fill_value=0)
    for k, letter in ORD_INV.items():
        s0_rows.append((f'  {letter}', int(vc[k]), f"{vc[k]/n_ds*100:.1f}%"))

s0_rows.append(('', '', ''))
s0_rows.append(('SECTION 0f: Prompt Count Distributions', '', ''))
for cond, col in [('Natural', 'nat_count'), ('Multiturn', 'mt_count')]:
    vals = df[col].dropna()
    q1, q3 = vals.quantile(0.25), vals.quantile(0.75)
    s0_rows.append((f'{cond} – Overall',
                    f"{vals.mean():.1f} / {vals.median():.1f} / {q1:.1f}–{q3:.1f} / {vals.min():.0f} / {vals.max():.0f}", ''))

df0 = pd.DataFrame(s0_rows, columns=['Metric','Value','Detail'])
df0.to_excel(writer, sheet_name='SUPP - Descriptives', index=False)
style_sheet(writer.sheets['SUPP - Descriptives'], header_rows=[1])
print("  Step 0 complete.")

# =====================================================================
# STEP 2: PRIMARY ANALYSIS
# =====================================================================
print("Running Step 2...")
m2a = compute_metrics(df['nat_lo'], df['nat_hi'], df['gold_ord'])
m2b = compute_metrics(df['mt_lo'],  df['mt_hi'],  df['gold_ord'])
conf2a = make_confusion(df['nat_lo'], df['nat_hi'], df['gold_ord'])
conf2b = make_confusion(df['mt_lo'],  df['mt_hi'],  df['gold_ord'])

sn2 = 'PRIMARY - Performance'
r = 0
pd.DataFrame([['STEP 2: PRIMARY ANALYSIS – GPT Conditions vs. Nurse Triage and Clinician-Adjudicated']]).to_excel(
    writer, sheet_name=sn2, startrow=r, index=False, header=False); r += 1
pd.DataFrame([[
    "Step 2 is the primary analysis: GPT-assisted triage decisions (Natural and Multiturn conditions) are evaluated "
    "against both Nurse Triage (Section 2A) and Clinician-Adjudicated (Section 2B) as independent reference standards. "
    "Range values use minimum-distance scoring (distance=0 if the reference level falls within the decision range). "
    "Cohen's kappa uses oracle assignment for range rows. "
    "For Section 2B, the 17 ambiguous ClinAdj values (A/B, B/C, C/D) use the conservative (upper urgency) bound as the reference."
]]).to_excel(writer, sheet_name=sn2, startrow=r, index=False, header=False); r += 2

pd.DataFrame([['2A: GPT Conditions vs. Nurse Triage']]).to_excel(
    writer, sheet_name=sn2, startrow=r, index=False, header=False); r += 1

pd.DataFrame({
    'Metric': [x[0] for x in metrics_to_rows(m2a)],
    'Natural Condition': [x[1] for x in metrics_to_rows(m2a)],
    'Multiturn Condition': [x[1] for x in metrics_to_rows(m2b)],
}).to_excel(writer, sheet_name=sn2, startrow=r, index=False); r += len(metrics_to_rows(m2a)) + 2

for label, conf in [('Natural', conf2a), ('Multiturn', conf2b)]:
    pd.DataFrame([[f'{label} – Confusion Matrix (Row=Decision oracle, Col=Nurse Triage)']]).to_excel(
        writer, sheet_name=sn2, startrow=r, index=False, header=False); r += 1
    conf.to_excel(writer, sheet_name=sn2, startrow=r); r += len(conf) + 2

print(f"  Step 2A complete. Nat agree={m2a['agree_pct']:.1f}%, MT agree={m2b['agree_pct']:.1f}%")

# =====================================================================
# STEP 3: CLINICIAN BENCHMARK
# =====================================================================
print("Running Step 3...")
m3a = compute_metrics(df['cc_lo'], df['cc_hi'], df['gold_ord'])
conf3 = make_confusion(df['cc_lo'], df['cc_hi'], df['gold_ord'])

sn3 = 'SUPP - ClinAdj vs Nurse'
r3 = 0
pd.DataFrame([['STEP 3: CLINICIAN CONSENSUS VS. NURSE TRIAGE (Physician Benchmark)']]).to_excel(
    writer, sheet_name=sn3, startrow=r3, index=False, header=False); r3 += 1
pd.DataFrame([[
    "Step 3 establishes the physician performance benchmark. The single Clinician-Adjudicated column applies to both conditions. "
    "Range consensus values (e.g., C/D, B/C, A/B — found in 17 rows) are treated as ranges: "
    "if the Nurse Triage level falls within the consensus range, it is counted as agreement (distance=0)."
]]).to_excel(writer, sheet_name=sn3, startrow=r3, index=False, header=False); r3 += 2
pd.DataFrame({
    'Metric': [x[0] for x in metrics_to_rows(m3a)],
    'Clinician-Adjudicated vs. Nurse Triage': [x[1] for x in metrics_to_rows(m3a)],
}).to_excel(writer, sheet_name=sn3, startrow=r3, index=False); r3 += len(metrics_to_rows(m3a)) + 2
pd.DataFrame([['Confusion Matrix: Clinician-Adjudicated (oracle) vs. Nurse Triage']]).to_excel(
    writer, sheet_name=sn3, startrow=r3, index=False, header=False); r3 += 1
conf3.to_excel(writer, sheet_name=sn3, startrow=r3); r3 += len(conf3) + 3

# --- Concordance & disagreement-direction analysis (ClinAdj vs. Nurse Triage) ---
_ccv = df['cc_cons_ord'].values.astype(float); _nuv = df['gold_ord'].values.astype(float)
_vm  = ~np.isnan(_ccv) & ~np.isnan(_nuv)
_cc  = _ccv[_vm].astype(int); _nu = _nuv[_vm].astype(int); _N = len(_cc)
_conc    = int((_cc == _nu).sum())
_conc_hi = int(np.isin(_nu[_cc == _nu], [2, 3]).sum())
_disc    = _cc != _nu; _nd = int(_disc.sum())
_nu_hi   = int((_nu > _cc).sum()); _cc_hi = int((_cc > _nu).sum())
_d3      = int((np.abs(_cc - _nu) == 3).sum())
_ue      = (np.isin(_cc, [2, 3]) | np.isin(_nu, [2, 3])) & _disc
_nue     = int(_ue.sum()); _nuhi_ue = int((_nu[_ue] > _cc[_ue]).sum())
def _ci2(k, n):
    if n == 0: return 'N/A'
    c = binomtest(k, n, 0.5).proportion_ci(0.95, method='exact')
    return f"{k}/{n} = {k/n*100:.1f}% (95% CI {c.low*100:.1f}-{c.high*100:.1f}%)"
_pdir = binomtest(_nu_hi, _nd, 0.5).pvalue if _nd else float('nan')
_pue  = binomtest(_nuhi_ue, _nue, 0.5).pvalue if _nue else float('nan')

pd.DataFrame([['Concordance & Disagreement Direction (Clinician-Adjudicated vs. Nurse Triage)']]).to_excel(
    writer, sheet_name=sn3, startrow=r3, index=False, header=False); r3 += 1
_dir_rows = [
    ('Concordant (exact agreement)',                          _ci2(_conc, _N)),
    ('  of which higher-acuity (C/D)',                        _ci2(_conc_hi, _conc)),
    ('Maximal disagreement (differ by 3 categories)',         _ci2(_d3, _N)),
    ('Discordant cases (N)',                                   f"{_nd}"),
    ('  Nurse higher acuity than clinician',                  f"{_ci2(_nu_hi, _nd)}  (binomial p={_pdir:.3f})"),
    ('  Clinician higher acuity than nurse',                  _ci2(_cc_hi, _nd)),
    ('  Nurse higher among urgent/emergent (ClinAdj or Nurse=C/D)', f"{_ci2(_nuhi_ue, _nue)}  (binomial p={_pue:.3f})"),
]
pd.DataFrame(_dir_rows, columns=['Metric', 'Value']).to_excel(
    writer, sheet_name=sn3, startrow=r3, index=False); r3 += len(_dir_rows) + 2

style_sheet(writer.sheets[sn3], title_row=1, header_rows=[4])
print(f"  Step 3 complete. ClinAdj agree={m3a['agree_pct']:.1f}%, kappa={m3a['kappa']:.3f}")

# =====================================================================
# STEP 3b: LLM CONDITIONS VS. CLINICIAN CONSENSUS AS REFERENCE
# =====================================================================
print("Running Step 3b...")

# --- Descriptive metrics ---
m3b_nat = compute_metrics(df['nat_lo'], df['nat_hi'], df['cc_cons_ord'])
m3b_mt  = compute_metrics(df['mt_lo'],  df['mt_hi'],  df['cc_cons_ord'])
conf3b_nat = make_confusion(df['nat_lo'], df['nat_hi'], df['cc_cons_ord'])
conf3b_mt  = make_confusion(df['mt_lo'],  df['mt_hi'],  df['cc_cons_ord'])

# --- Per-row arrays for statistical tests ---
nat_dist_cc   = row_dist(df['nat_lo'],   df['nat_hi'],  df['cc_cons_ord'])
mt_dist_cc    = row_dist(df['mt_lo'],    df['mt_hi'],   df['cc_cons_ord'])
nat_signed_cc = row_signed(df['nat_lo'], df['nat_hi'],  df['cc_cons_ord'])
mt_signed_cc  = row_signed(df['mt_lo'],  df['mt_hi'],   df['cc_cons_ord'])

# Paired mask (needs valid nat, mt, and ClinAdj gold)
mask3b = (~df['nat_lo'].isna()) & (~df['mt_lo'].isna()) & (~df['cc_cons_ord'].isna())
n3b    = int(mask3b.sum())
nd_cc  = nat_dist_cc[mask3b]
md_cc  = mt_dist_cc[mask3b]
ns_cc  = nat_signed_cc[mask3b]
ms_cc  = mt_signed_cc[mask3b]
na_cc  = (nd_cc == 0)
ma_cc  = (md_cc == 0)
nu_cc  = (ns_cc < 0)
mu_cc  = (ms_cc < 0)

# Wilcoxon signed-rank (distance)
try:
    w3b_stat, w3b_p = wilcoxon(nd_cc, md_cc, zero_method='wilcox')
    # Effect size r = |Z| / sqrt(N pairs)
    _z3b = wilcoxon(nd_cc, md_cc, zero_method='wilcox', correction=True, method='approx').zstatistic
    w3b_r = abs(_z3b) / np.sqrt(n3b)
except Exception:
    w3b_stat, w3b_p, w3b_r = np.nan, np.nan, np.nan

# McNemar (agreement and under-triage)
mcn3b_agree_s, mcn3b_agree_p = mcnemar_test(na_cc, ma_cc)
mcn3b_under_s, mcn3b_under_p = mcnemar_test(nu_cc, mu_cc)
d3b_eff = cohens_d_paired(nd_cc, md_cc)

# Binomial test for direction asymmetry vs ClinAdj gold
from scipy.stats import binomtest as _btest
def _asym(signed_arr):
    s = np.array(signed_arr, float)
    n_e = int((s != 0).sum()); n_u = int((s < 0).sum())
    if n_e == 0: return n_e, n_u, int((s>0).sum()), np.nan, np.nan, np.nan, 'N/A'
    res = _btest(n_u, n_e, p=0.5, alternative='two-sided')
    ci  = res.proportion_ci(confidence_level=0.95)
    sig = 'Yes' if res.pvalue < 0.05 else 'No'
    return n_e, n_u, int((s>0).sum()), n_u/n_e, res.pvalue, (ci.low, ci.high), sig

asym3b_nat = _asym(ns_cc)
asym3b_mt  = _asym(ms_cc)

# ---- Write Section 2B into Step 2 sheet ----
pd.DataFrame([['2B: GPT Conditions vs. Clinician-Adjudicated']]).to_excel(
    writer, sheet_name=sn2, startrow=r, index=False, header=False); r += 1
pd.DataFrame([[
    "Section 2B evaluates GPT-assisted triage (Natural and Multiturn) against Clinician-Adjudicated as the reference. "
    "Distance = 0 when the LLM decision range includes the ClinAdj reference value. "
    "Under-triage here means the LLM assigned lower urgency than physician consensus. "
    "Subsection i: Descriptive metrics. Subsection ii: Natural vs. Multiturn paired comparison (ClinAdj as reference). "
    "Subsection iii: Direction-of-error asymmetry (binomial test, H0: P(under|error)=0.5)."
]]).to_excel(writer, sheet_name=sn2, startrow=r, index=False, header=False); r += 2

# --- 2B-i: Descriptive metrics vs. Clinician-Adjudicated ---
pd.DataFrame([['2B-i: Descriptive Metrics vs. Clinician-Adjudicated']]).to_excel(
    writer, sheet_name=sn2, startrow=r, index=False, header=False); r += 1
pd.DataFrame({
    'Metric': [x[0] for x in metrics_to_rows(m3b_nat)],
    'Natural vs. Clinician-Adjudicated': [x[1] for x in metrics_to_rows(m3b_nat)],
    'Multiturn vs. Clinician-Adjudicated': [x[1] for x in metrics_to_rows(m3b_mt)],
}).to_excel(writer, sheet_name=sn2, startrow=r, index=False)
r += len(metrics_to_rows(m3b_nat)) + 2

for label, conf in [('Natural', conf3b_nat), ('Multiturn', conf3b_mt)]:
    pd.DataFrame([[f'{label} – Confusion Matrix (Row=Decision oracle, Col=Clinician-Adjudicated)']]).to_excel(
        writer, sheet_name=sn2, startrow=r, index=False, header=False); r += 1
    conf.to_excel(writer, sheet_name=sn2, startrow=r); r += len(conf) + 2

# --- 2B-ii: Paired comparison Natural vs. Multiturn (ClinAdj as reference) ---
pd.DataFrame([['2B-ii: Paired Comparison – Natural vs. Multiturn (Clinician-Adjudicated as Reference)']]).to_excel(
    writer, sheet_name=sn2, startrow=r, index=False, header=False); r += 1
rows2b_paired = [
    ('N pairs (valid for all three: Nat, MT, ClinAdj)',   n3b),
    ('Natural – Mean Distance from ClinAdj (SD)',
     f"{nd_cc.mean():.3f} ({nd_cc.std(ddof=1):.3f})"),
    ('Multiturn – Mean Distance from ClinAdj (SD)',
     f"{md_cc.mean():.3f} ({md_cc.std(ddof=1):.3f})"),
    ('Natural – Exact Agreement with ClinAdj (%)',   f"{na_cc.mean()*100:.1f}%"),
    ('Multiturn – Exact Agreement with ClinAdj (%)', f"{ma_cc.mean()*100:.1f}%"),
    ('Natural – Under-triage vs. ClinAdj (%)',   f"{nu_cc.mean()*100:.1f}%"),
    ('Multiturn – Under-triage vs. ClinAdj (%)', f"{mu_cc.mean()*100:.1f}%"),
    ('', ''),
    ('Wilcoxon signed-rank (distance)',     f"{w3b_stat:.3f}" if not np.isnan(w3b_stat) else 'N/A'),
    ('Wilcoxon p-value',                    f"{w3b_p:.4f}" if not np.isnan(w3b_p) else 'N/A'),
    ('Wilcoxon significant (α=0.05)?',      'Yes ***' if (not np.isnan(w3b_p) and w3b_p < 0.05) else 'No'),
    ('Effect size r (Wilcoxon)',            f"{w3b_r:.3f}" if not np.isnan(w3b_r) else 'N/A'),
    ('', ''),
    ('McNemar χ² (agreement)',              f"{mcn3b_agree_s:.3f}" if not np.isnan(mcn3b_agree_s) else 'N/A'),
    ('McNemar p-value (agreement)',         f"{mcn3b_agree_p:.4f}" if not np.isnan(mcn3b_agree_p) else 'N/A'),
    ('Agreement significant (α=0.05)?',    'Yes ***' if (not np.isnan(mcn3b_agree_p) and mcn3b_agree_p < 0.05) else 'No'),
    ('', ''),
    ('McNemar χ² (under-triage)',           f"{mcn3b_under_s:.3f}" if not np.isnan(mcn3b_under_s) else 'N/A'),
    ('McNemar p-value (under-triage)',      f"{mcn3b_under_p:.4f}" if not np.isnan(mcn3b_under_p) else 'N/A'),
    ('Under-triage significant (α=0.05)?', 'Yes ***' if (not np.isnan(mcn3b_under_p) and mcn3b_under_p < 0.05) else 'No'),
    ('', ''),
    ("Cohen's d (distance, Nat–MT)",        f"{d3b_eff:.3f}" if not np.isnan(d3b_eff) else 'N/A'),
]
pd.DataFrame(rows2b_paired, columns=['Metric', 'Value']).to_excel(
    writer, sheet_name=sn2, startrow=r, index=False)
r += len(rows2b_paired) + 2

# --- 2B-iii: Direction asymmetry (binomial test) ---
pd.DataFrame([['2B-iii: Direction-of-Error Asymmetry vs. Clinician-Adjudicated (H0: P(under|error) = 0.5)']]).to_excel(
    writer, sheet_name=sn2, startrow=r, index=False, header=False); r += 1

asym_cols3b = ['Condition', 'N Errors', 'N Under', 'N Over',
               '% Under Among Errors', '95% CI', 'Binomial p-value', 'Reject H0 (p<0.05)?']
asym_data3b = []
for lbl, (n_e, n_u, n_o, p_u, pv, ci, sig) in [
    ('Natural vs. Clinician-Adjudicated',   asym3b_nat),
    ('Multiturn vs. Clinician-Adjudicated', asym3b_mt),
]:
    asym_data3b.append([
        lbl, n_e, n_u, n_o,
        f"{p_u*100:.1f}%" if not isinstance(p_u, float) or not np.isnan(p_u) else 'N/A',
        f"({ci[0]*100:.1f}%, {ci[1]*100:.1f}%)" if not isinstance(ci, str) else 'N/A',
        f"{pv:.4f}" if not np.isnan(pv) else 'N/A',
        sig,
    ])
pd.DataFrame(asym_data3b, columns=asym_cols3b).to_excel(
    writer, sheet_name=sn2, startrow=r, index=False)
r += len(asym_data3b) + 2

pd.DataFrame([[
    f"NOTE: 17 ambiguous ClinAdj rows use conservative upper bound as reference value. "
    f"Under-triage = LLM decision urgency < physician consensus urgency. "
    f"Natural: {asym3b_nat[1]}/{asym3b_nat[0]} errors are under-triage "
    f"({'significant' if asym3b_nat[4]<0.05 else 'not significant'}, p={asym3b_nat[4]:.4f}). "
    f"Multiturn: {asym3b_mt[1]}/{asym3b_mt[0]} errors are under-triage "
    f"({'significant' if asym3b_mt[4]<0.05 else 'not significant'}, p={asym3b_mt[4]:.4f})."
]]).to_excel(writer, sheet_name=sn2, startrow=r, index=False, header=False)

style_sheet(writer.sheets[sn2], title_row=1, header_rows=[4])
print(f"  Step 2 complete. Nat/NT agree={m2a['agree_pct']:.1f}%, MT/NT agree={m2b['agree_pct']:.1f}%, "
      f"Nat/ClinAdj agree={m3b_nat['agree_pct']:.1f}%, MT/ClinAdj agree={m3b_mt['agree_pct']:.1f}%")


# =====================================================================
# STEP 4: SENSITIVITY ANALYSIS
# =====================================================================
print("Running Step 4...")
sn4 = 'SECONDARY - Sensitivity'
r4 = 0
pd.DataFrame([['STEP 4: SENSITIVITY ANALYSIS – Conservative vs. Aggressive Resolution']]).to_excel(
    writer, sheet_name=sn4, startrow=r4, index=False, header=False); r4 += 1
pd.DataFrame([[
    "Step 4a quantifies inter-rater agreement between Primary Grader and Reviewer. "
    "Steps 4b/c apply Conservative (higher urgency wins) and Aggressive (lower urgency wins) resolution. "
    "Grader 5 rows are flagged (same person graded and reviewed)."
]]).to_excel(writer, sheet_name=sn4, startrow=r4, index=False, header=False); r4 += 2

def inter_rater(lo1, hi1, lo2, hi2, label, df_sub):
    m = compute_metrics(df_sub[lo1], df_sub[hi1], df_sub[hi2])
    return [
        (f'{label} – N', m['n']),
        (f'{label} – Exact Agreement (%)',
         f"{m['agree_pct']:.1f}% (95%CI: {m['agree_ci_lo']:.1f}–{m['agree_ci_hi']:.1f}%)"),
        (f'{label} – Weighted Kappa (95%CI)',
         f"{m['kappa']:.3f} ({m['kappa_lo']:.3f}, {m['kappa_hi']:.3f})" if not np.isnan(m['kappa']) else 'N/A'),
        (f'{label} – Disagreement 1 step (%)', f"{m['dist_pcts'][1]:.1f}%"),
        (f'{label} – Disagreement 2 steps (%)', f"{m['dist_pcts'][2]:.1f}%"),
        (f'{label} – Disagreement 3 steps (%)', f"{m['dist_pcts'][3]:.1f}%"),
    ]

df_no5 = df[df['grader'] != 5]
pd.DataFrame([[f"Note: {len(df[df['grader']==5])} rows = Grader 5 (self-review)"]]).to_excel(
    writer, sheet_name=sn4, startrow=r4, index=False, header=False); r4 += 1
ir_rows = (inter_rater('nat_lo','nat_hi','nat_rev_lo','nat_rev_hi','Natural (All)', df) +
           inter_rater('mt_lo', 'mt_hi', 'mt_rev_lo', 'mt_rev_hi', 'Multiturn (All)', df) +
           inter_rater('nat_lo','nat_hi','nat_rev_lo','nat_rev_hi','Natural (excl. Grader 5)', df_no5) +
           inter_rater('mt_lo', 'mt_hi', 'mt_rev_lo', 'mt_rev_hi', 'Multiturn (excl. Grader 5)', df_no5))
pd.DataFrame(ir_rows, columns=['Metric','Value']).to_excel(writer, sheet_name=sn4, startrow=r4, index=False); r4 += len(ir_rows)+2

m4_nat_cons = compute_metrics(df['nat_resolved_cons'], df['nat_resolved_cons'], df['gold_ord'])
m4_nat_agg  = compute_metrics(df['nat_resolved_agg'],  df['nat_resolved_agg'],  df['gold_ord'])
m4_mt_cons  = compute_metrics(df['mt_resolved_cons'],  df['mt_resolved_cons'],  df['gold_ord'])
m4_mt_agg   = compute_metrics(df['mt_resolved_agg'],   df['mt_resolved_agg'],   df['gold_ord'])

rp = metrics_to_rows(m2a); rc = metrics_to_rows(m4_nat_cons)
ra = metrics_to_rows(m4_nat_agg); rmp= metrics_to_rows(m2b)
rmc= metrics_to_rows(m4_mt_cons); rma= metrics_to_rows(m4_mt_agg)

pd.DataFrame([['4d. Side-by-Side: Primary | Conservative | Aggressive']]).to_excel(
    writer, sheet_name=sn4, startrow=r4, index=False, header=False); r4 += 1
pd.DataFrame({
    'Metric': [x[0] for x in rp],
    'Natural – Primary':        [x[1] for x in rp],
    'Natural – Conservative':   [x[1] for x in rc],
    'Natural – Aggressive':     [x[1] for x in ra],
    'Multiturn – Primary':      [x[1] for x in rmp],
    'Multiturn – Conservative': [x[1] for x in rmc],
    'Multiturn – Aggressive':   [x[1] for x in rma],
}).to_excel(writer, sheet_name=sn4, startrow=r4, index=False)
style_sheet(writer.sheets[sn4], title_row=1, header_rows=[4])
print("  Step 4 complete.")

# =====================================================================
# STEP 5: SUBGROUP BY DATASET
# =====================================================================
print("Running Step 5...")
sn5 = 'SECONDARY - By Dataset'
r5 = 0
pd.DataFrame([['STEP 5: SUBGROUP ANALYSIS BY DATASET']]).to_excel(
    writer, sheet_name=sn5, startrow=r5, index=False, header=False); r5 += 1
pd.DataFrame([[
    "Step 5 stratifies all primary metrics by dataset. ⚠ flags datasets differing from overall by "
    ">10 pp in exact agreement or >0.5 in mean distance."
]]).to_excel(writer, sheet_name=sn5, startrow=r5, index=False, header=False); r5 += 2

overall_nat_agree = m2a['agree_pct']; overall_nat_dist = m2a['mean_dist']
overall_mt_agree  = m2b['agree_pct']; overall_mt_dist  = m2b['mean_dist']
overall_cc_agree  = m3a['agree_pct']; overall_cc_dist  = m3a['mean_dist']

summary5 = []
for ds in [1,2,3]:
    sub = df[df['dataset']==ds]
    mn = compute_metrics(sub['nat_lo'], sub['nat_hi'], sub['gold_ord'])
    mm = compute_metrics(sub['mt_lo'],  sub['mt_hi'],  sub['gold_ord'])
    mc = compute_metrics(sub['cc_lo'],  sub['cc_hi'],  sub['gold_ord'])
    fn = '⚠' if abs(mn.get('agree_pct',0)-overall_nat_agree)>10 or abs(mn.get('mean_dist',0)-overall_nat_dist)>0.5 else ''
    fm = '⚠' if abs(mm.get('agree_pct',0)-overall_mt_agree)>10  or abs(mm.get('mean_dist',0)-overall_mt_dist)>0.5  else ''
    fc = '⚠' if abs(mc.get('agree_pct',0)-overall_cc_agree)>10  or abs(mc.get('mean_dist',0)-overall_cc_dist)>0.5  else ''

    vc = sub['gold_ord'].value_counts().reindex([0,1,2,3], fill_value=0)
    gs = {f'Nurse {ORD_INV[k]}': int(vc[k]) for k in [0,1,2,3]}
    pd.DataFrame([[f'Dataset {ds} (N={mn["n"]}) Nurse Triage distribution: {gs}']]).to_excel(
        writer, sheet_name=sn5, startrow=r5, index=False, header=False); r5 += 1

    all_rows = []
    for m, lbl in [(mn,f'Natural{fn}'),(mm,f'Multiturn{fm}'),(mc,f'Clin.Consensus{fc}')]:
        all_rows.extend([(f'{lbl}: {x[0]}', x[1]) for x in metrics_to_rows(m)])
    pd.DataFrame(all_rows, columns=['Metric','Value']).to_excel(
        writer, sheet_name=sn5, startrow=r5, index=False); r5 += len(all_rows)+2

    summary5.append({
        'Dataset': ds, 'N': mn['n'],
        'Nat Agree%': f"{mn.get('agree_pct',0):.1f}%",
        'Nat Kappa':  f"{mn.get('kappa',np.nan):.3f}" if not np.isnan(mn.get('kappa',np.nan)) else 'N/A',
        'Nat MeanDist': f"{mn.get('mean_dist',0):.3f}",
        'Nat Under%': f"{mn.get('under_pct',0):.1f}%",
        'MT Agree%':  f"{mm.get('agree_pct',0):.1f}%",
        'MT Kappa':   f"{mm.get('kappa',np.nan):.3f}" if not np.isnan(mm.get('kappa',np.nan)) else 'N/A',
        'MT MeanDist':f"{mm.get('mean_dist',0):.3f}",
        'MT Under%':  f"{mm.get('under_pct',0):.1f}%",
        'ClinAdj Agree%':  f"{mc.get('agree_pct',0):.1f}%",
        'ClinAdj Kappa':   f"{mc.get('kappa',np.nan):.3f}" if not np.isnan(mc.get('kappa',np.nan)) else 'N/A',
        'ClinAdj MeanDist':f"{mc.get('mean_dist',0):.3f}",
        'ClinAdj Under%':  f"{mc.get('under_pct',0):.1f}%",
        'Flag(Nat)': fn, 'Flag(MT)': fm, 'Flag(ClinAdj)': fc,
    })

pd.DataFrame([['5f. Cross-Dataset Summary (⚠ = meaningfully different from overall)']]).to_excel(
    writer, sheet_name=sn5, startrow=r5, index=False, header=False); r5 += 1
pd.DataFrame(summary5).to_excel(writer, sheet_name=sn5, startrow=r5, index=False)
r5 += len(summary5) + 2

# --- 5g. Cross-Dataset Generalization Tests (synthetic vignettes vs. real-world) ---
from scipy.stats import kruskal as _kruskal, chi2_contingency as _chi2c
pd.DataFrame([['5g. Cross-Dataset Generalization: synthetic vignettes (Dataset 1) vs. real-world (Datasets 2-3)']]).to_excel(
    writer, sheet_name=sn5, startrow=r5, index=False, header=False); r5 += 1
pd.DataFrame([[
    "Exact agreement and ordinal distance compared across the three datasets (chi-square on agreement; "
    "Kruskal-Wallis on distance) and as a pre-planned contrast of vignettes (Dataset 1) vs. pooled "
    "real-world encounters (Datasets 2-3; chi-square on agreement, Mann-Whitney U on distance)."
]]).to_excel(writer, sheet_name=sn5, startrow=r5, index=False, header=False); r5 += 2

def _gen_row(cond_lbl, lo, hi, ref_lbl, ref):
    dd = {ds: np.array(row_dist(df[df['dataset']==ds][lo], df[df['dataset']==ds][hi],
                                df[df['dataset']==ds][ref]), float) for ds in [1,2,3]}
    dd = {ds: v[~np.isnan(v)] for ds, v in dd.items()}
    ag = {ds: (dd[ds]==0).mean()*100 for ds in [1,2,3]}
    H, pH = _kruskal(dd[1], dd[2], dd[3])
    tab3 = [[int((dd[ds]==0).sum()), int((dd[ds]!=0).sum())] for ds in [1,2,3]]
    chi3, pc3, _, _ = _chi2c(tab3)
    rw = np.concatenate([dd[2], dd[3]])
    U, pU = mannwhitneyu(dd[1], rw, alternative='two-sided')
    tab2 = [[int((dd[1]==0).sum()), int((dd[1]!=0).sum())],
            [int((rw==0).sum()),    int((rw!=0).sum())]]
    chi2v, pc2, _, _ = _chi2c(tab2)
    return [f'{cond_lbl} vs. {ref_lbl}',
            f"{ag[1]:.1f}%", f"{ag[2]:.1f}%", f"{ag[3]:.1f}%", f"{(rw==0).mean()*100:.1f}%",
            f"{chi3:.2f}", f"{pc3:.4f}", f"{pH:.4f}", f"{chi2v:.2f}", f"{pc2:.4f}", f"{pU:.4f}"]

gen_rows = []
for cn, lo, hi in [('Natural','nat_lo','nat_hi'), ('Multiturn','mt_lo','mt_hi')]:
    for rn, rc in [('Nurse Triage','gold_ord'), ('Clinician-Adjudicated','cc_cons_ord')]:
        gen_rows.append(_gen_row(cn, lo, hi, rn, rc))
gen_cols = ['Comparison', 'Vignettes (n=39) agree', 'ED (n=76) agree', 'Nurse-line (n=140) agree',
            'Real-world (n=216) agree', 'Omnibus chi2 (agree)', 'Omnibus p (agree)',
            'Omnibus KW p (distance)', 'Vig-vs-RW chi2 (agree)', 'Vig-vs-RW p (agree)',
            'Vig-vs-RW MWU p (distance)']
pd.DataFrame(gen_rows, columns=gen_cols).to_excel(writer, sheet_name=sn5, startrow=r5, index=False)
r5 += len(gen_rows) + 2

style_sheet(writer.sheets[sn5], title_row=1, header_rows=[4])
print("  Step 5 complete.")

# =====================================================================
# STEP 5b: UNDER-/OVER-TRIAGE BY REFERENCE LEVEL, DATASET & COMPARATOR
# =====================================================================
print("Running Step 5b...")
sn5b = 'SECONDARY - Dataset x Level'
r5b = 0
pd.DataFrame([['STEP 5b: UNDER-/OVER-TRIAGE BY REFERENCE TRIAGE LEVEL, DATASET & COMPARATOR']]).to_excel(
    writer, sheet_name=sn5b, startrow=r5b, index=False, header=False); r5b += 1
pd.DataFrame([[
    "For each dataset and comparator, cases are grouped by the reference standard's triage level (A–D) and "
    "split into under-triage, correct, and over-triage. 'Reference Level' = the comparator's own reference: "
    "Nurse Triage for the 'vs. Nurse Triage' rows, Clinician-Adjudicated (conservative/upper bound for the 17 "
    "ambiguous ClinAdj values) for the 'vs. Clinician-Adjudicated' rows. Under-triage = LLM assigned lower urgency than "
    "the reference; over-triage = higher urgency. Note: level A cannot be under-triaged and level D cannot be "
    "over-triaged. Percentages are of cases at that reference level (Under+Correct+Over = 100%). 'All Datasets' "
    "pools all rows; Dataset 3 is the 'Nurse Triage' vignette set (distinct from the Nurse Triage reference)."
]]).to_excel(writer, sheet_name=sn5b, startrow=r5b, index=False, header=False); r5b += 2

def level_strat_counts(sub, dec_lo, dec_hi, ref_col, level_ord):
    sgn = np.array(row_signed(sub[dec_lo], sub[dec_hi], sub[ref_col]), float)
    ref = np.array(sub[ref_col], float)
    m   = (ref == level_ord) & ~np.isnan(sgn)
    s   = sgn[m]
    return len(s), int((s < 0).sum()), int((s == 0).sum()), int((s > 0).sum())

comparators_5b = [
    ('Natural vs. Nurse Triage',          'nat_lo', 'nat_hi', 'gold_ord'),
    ('Multiturn vs. Nurse Triage',        'mt_lo',  'mt_hi',  'gold_ord'),
    ('Natural vs. Clinician-Adjudicated',   'nat_lo', 'nat_hi', 'cc_cons_ord'),
    ('Multiturn vs. Clinician-Adjudicated', 'mt_lo',  'mt_hi',  'cc_cons_ord'),
]
ds_names_5b = {1: 'CAVs', 2: 'Emergency Department', 3: 'Nurse Triage vignettes'}
dataset_groups_5b = [('All Datasets', df)] + [
    (f'Dataset {ds} — {ds_names_5b[ds]}', df[df['dataset'] == ds]) for ds in [1, 2, 3]
]

def _pct5b(x, n):
    return f"{x/n*100:.1f}%" if n > 0 else '—'

rows_5b = []
for ds_lbl, sub in dataset_groups_5b:
    for comp_lbl, dlo, dhi, refc in comparators_5b:
        t_n = t_u = t_c = t_o = 0
        for k in [0, 1, 2, 3]:
            n, u, c, o = level_strat_counts(sub, dlo, dhi, refc, k)
            t_n += n; t_u += u; t_c += c; t_o += o
            rows_5b.append([ds_lbl, comp_lbl, ORD_INV[k], n,
                            u, _pct5b(u, n), c, _pct5b(c, n), o, _pct5b(o, n)])
        rows_5b.append([ds_lbl, comp_lbl, 'All A–D', t_n,
                        t_u, _pct5b(t_u, t_n), t_c, _pct5b(t_c, t_n), t_o, _pct5b(t_o, t_n)])

cols_5b = ['Dataset', 'Comparator', 'Reference Level', 'N at Level',
           'Under-triage (n)', 'Under-triage (%)', 'Correct (n)', 'Correct (%)',
           'Over-triage (n)', 'Over-triage (%)']
pd.DataFrame(rows_5b, columns=cols_5b).to_excel(writer, sheet_name=sn5b, startrow=r5b, index=False)
style_sheet(writer.sheets[sn5b], title_row=1, header_rows=[4])
print(f"  Step 5b complete. {len(rows_5b)} rows written.")

# =====================================================================
# STEP 6: NATURAL vs. MULTITURN COMPARISON
# =====================================================================
print("Running Step 6...")
sn6 = 'SECONDARY - Nat vs Multiturn'
r6 = 0
pd.DataFrame([['STEP 6: CONDITION COMPARISON – Natural vs. Multiturn']]).to_excel(
    writer, sheet_name=sn6, startrow=r6, index=False, header=False); r6 += 1
pd.DataFrame([[
    "Step 6 compares Natural vs. Multiturn on the same prompts using paired tests. "
    "Wilcoxon signed-rank is used for distance scores (effect size r = |Z|/sqrt(N pairs)); "
    "McNemar tests binary outcomes (agreement, under-triage). Bonferroni-adjusted p-values "
    "(raw x 2, capped at 1.0) correct the primary Overall comparison across the two reference "
    "standards; significance threshold alpha = 0.05 on adjusted p (equivalently alpha = 0.025 on raw p)."
]]).to_excel(writer, sheet_name=sn6, startrow=r6, index=False, header=False); r6 += 2

def condition_compare(df_sub, label='Overall', ref='gold_ord'):
    mask = (~df_sub['nat_lo'].isna()) & (~df_sub['mt_lo'].isna()) & (~df_sub[ref].isna())
    sub  = df_sub[mask]
    nat_d = np.array(row_dist(sub['nat_lo'], sub['nat_hi'], sub[ref]), float)
    mt_d  = np.array(row_dist(sub['mt_lo'],  sub['mt_hi'],  sub[ref]), float)
    nat_s = np.array(row_signed(sub['nat_lo'], sub['nat_hi'], sub[ref]), float)
    mt_s  = np.array(row_signed(sub['mt_lo'],  sub['mt_hi'],  sub[ref]), float)
    nat_a = (nat_d == 0); mt_a = (mt_d == 0)
    nat_u = (nat_s < 0);  mt_u = (mt_s < 0)
    n = len(nat_d)
    try:
        wstat, wpval = wilcoxon(nat_d, mt_d, zero_method='wilcox')
        # Effect size r = |Z| / sqrt(N pairs)
        _zw = wilcoxon(nat_d, mt_d, zero_method='wilcox', correction=True, method='approx').zstatistic
        r_eff = abs(_zw) / np.sqrt(n) if n > 0 else np.nan
    except Exception:
        wstat, wpval, r_eff = np.nan, np.nan, np.nan
    mcn_agree_s, mcn_agree_p = mcnemar_test(nat_a, mt_a)
    mcn_under_s, mcn_under_p = mcnemar_test(nat_u, mt_u)
    d_eff = cohens_d_paired(nat_d, mt_d)
    def _badj(p):  # Bonferroni adjustment across the two reference standards (x2, capped at 1)
        return 'N/A' if np.isnan(p) else f"{min(p*2, 1.0):.4f}"
    return [
        ('Label', label), ('N pairs', n),
        ('Natural – Mean Distance (SD)', f"{nat_d.mean():.3f} ({nat_d.std(ddof=1):.3f})"),
        ('Multiturn – Mean Distance (SD)', f"{mt_d.mean():.3f} ({mt_d.std(ddof=1):.3f})"),
        ('Natural – Exact Agreement (%)', f"{nat_a.mean()*100:.1f}%"),
        ('Multiturn – Exact Agreement (%)', f"{mt_a.mean()*100:.1f}%"),
        ('Natural – Under-triage (%)', f"{nat_u.mean()*100:.1f}%"),
        ('Multiturn – Under-triage (%)', f"{mt_u.mean()*100:.1f}%"),
        ('Wilcoxon statistic', f"{wstat:.3f}" if not np.isnan(wstat) else 'N/A'),
        ('Wilcoxon p-value (raw)', f"{wpval:.4f}" if not np.isnan(wpval) else 'N/A'),
        ('Wilcoxon p-value (Bonferroni-adj.)', _badj(wpval)),
        ('Effect size r (Wilcoxon)', f"{r_eff:.3f}" if not np.isnan(r_eff) else 'N/A'),
        ("McNemar χ² (agreement)", f"{mcn_agree_s:.3f}" if not np.isnan(mcn_agree_s) else 'N/A'),
        ("McNemar p-value (agreement, raw)", f"{mcn_agree_p:.4f}" if not np.isnan(mcn_agree_p) else 'N/A'),
        ("McNemar p-value (agreement, Bonferroni-adj.)", _badj(mcn_agree_p)),
        ("McNemar χ² (under-triage)", f"{mcn_under_s:.3f}" if not np.isnan(mcn_under_s) else 'N/A'),
        ("McNemar p-value (under-triage, raw)", f"{mcn_under_p:.4f}" if not np.isnan(mcn_under_p) else 'N/A'),
        ("McNemar p-value (under-triage, Bonferroni-adj.)", _badj(mcn_under_p)),
        ("Cohen's d (distance)", f"{d_eff:.3f}" if not np.isnan(d_eff) else 'N/A'),
    ]

for ref_lbl, ref_col in [('vs. Nurse Triage', 'gold_ord'), ('vs. Clinician-Adjudicated', 'cc_cons_ord')]:
    pd.DataFrame([[f'Natural vs. Multiturn — {ref_lbl}']]).to_excel(
        writer, sheet_name=sn6, startrow=r6, index=False, header=False); r6 += 2
    for sub_df, lbl in [(df, 'Overall')] + [(df[df['dataset']==ds], f'Dataset {ds}') for ds in [1,2,3]]:
        rows6 = condition_compare(sub_df, lbl, ref=ref_col)
        pd.DataFrame([[lbl]]).to_excel(writer, sheet_name=sn6, startrow=r6, index=False, header=False); r6 += 1
        pd.DataFrame(rows6, columns=['Metric','Value']).to_excel(writer, sheet_name=sn6, startrow=r6, index=False)
        r6 += len(rows6) + 2
style_sheet(writer.sheets[sn6], title_row=1, header_rows=[4])
print("  Step 6 complete.")


# =====================================================================
# STEP 7: SAFETY-CRITICAL ERROR ANALYSIS
# =====================================================================
print("Running Step 7...")
sn7 = 'PRIMARY - Safety-Critical'
r7 = 0
pd.DataFrame([['STEP 7: SAFETY-CRITICAL ERROR ANALYSIS']]).to_excel(
    writer, sheet_name=sn7, startrow=r7, index=False, header=False); r7 += 1
pd.DataFrame([[
    "Step 7 identifies cases with distance ≥ 2 in either condition (moderate/severe disagreement). "
    "Section 7a: vs. Nurse Triage. Section 7b: vs. Clinician-Adjudicated. "
    "Under-triage cases with distance ≥ 2 are flagged 'SAFETY-CRITICAL'."
]]).to_excel(writer, sheet_name=sn7, startrow=r7, index=False, header=False); r7 += 2

df7 = df.copy()
df7['nat_dir'] = np.where(df7['nat_signed'] < 0, 'Under', np.where(df7['nat_signed'] > 0, 'Over', 'Exact'))
df7['mt_dir']  = np.where(df7['mt_signed']  < 0, 'Under', np.where(df7['mt_signed']  > 0, 'Over', 'Exact'))
df7['safety_crit'] = (((df7['nat_dist'] >= 2) & (df7['nat_dir'] == 'Under')) |
                      ((df7['mt_dist']  >= 2) & (df7['mt_dir']  == 'Under')))

sev_either = df7[(df7['nat_dist'] >= 2) | (df7['mt_dist'] >= 2)].copy()
out7 = sev_either[['dataset','prompt','grader','complaint','gold_raw',
                    'nat_triage_raw','nat_dist','nat_dir',
                    'mt_triage_raw','mt_dist','mt_dir','safety_crit']].copy()
out7.columns = ['Dataset','Prompt','Grader','Chief Complaint','Nurse Triage',
                'Natural Decision','Nat Dist','Nat Dir',
                'Multiturn Decision','MT Dist','MT Dir','Safety-Critical']
out7['Safety-Critical'] = out7['Safety-Critical'].map({True:'*** SAFETY-CRITICAL', False:''})
pd.DataFrame([[f'7a. Cases with Distance ≥ 2 vs. Nurse Triage (N={len(out7)})']]).to_excel(
    writer, sheet_name=sn7, startrow=r7, index=False, header=False); r7 += 1
out7.to_excel(writer, sheet_name=sn7, startrow=r7, index=False); r7 += len(out7)+2

# --- 7b: Safety cases vs. Clinician-Adjudicated ---
nat_dist_cc_arr   = np.array(nat_dist_cc,   float)
mt_dist_cc_arr    = np.array(mt_dist_cc,    float)
nat_signed_cc_arr = np.array(nat_signed_cc, float)
mt_signed_cc_arr  = np.array(mt_signed_cc,  float)
df7['nat_dist_cc']  = nat_dist_cc_arr
df7['mt_dist_cc']   = mt_dist_cc_arr
df7['nat_dir_cc']   = np.where(nat_signed_cc_arr < 0, 'Under', np.where(nat_signed_cc_arr > 0, 'Over', 'Exact'))
df7['mt_dir_cc']    = np.where(mt_signed_cc_arr  < 0, 'Under', np.where(mt_signed_cc_arr  > 0, 'Over', 'Exact'))
df7['sc_cc'] = (((df7['nat_dist_cc'] >= 2) & (df7['nat_dir_cc'] == 'Under')) |
                ((df7['mt_dist_cc']  >= 2) & (df7['mt_dir_cc']  == 'Under')))
sev_cc = df7[(df7['nat_dist_cc'] >= 2) | (df7['mt_dist_cc'] >= 2)].copy()
out7b = sev_cc[['dataset','prompt','grader','complaint','cc_raw',
                'nat_triage_raw','nat_dist_cc','nat_dir_cc',
                'mt_triage_raw','mt_dist_cc','mt_dir_cc','sc_cc']].copy()
out7b.columns = ['Dataset','Prompt','Grader','Chief Complaint','Clinician-Adjudicated',
                 'Natural Decision','Nat Dist (ClinAdj)','Nat Dir (ClinAdj)',
                 'Multiturn Decision','MT Dist (ClinAdj)','MT Dir (ClinAdj)','Safety-Critical vs. ClinAdj']
out7b['Safety-Critical vs. ClinAdj'] = out7b['Safety-Critical vs. ClinAdj'].map({True:'*** SAFETY-CRITICAL', False:''})
pd.DataFrame([[f'7b. Cases with Distance ≥ 2 vs. Clinician-Adjudicated (N={len(out7b)})']]).to_excel(
    writer, sheet_name=sn7, startrow=r7, index=False, header=False); r7 += 1
out7b.to_excel(writer, sheet_name=sn7, startrow=r7, index=False); r7 += len(out7b)+2

sc_cases    = df7[df7['safety_crit']]
sc_cc_cases = df7[df7['sc_cc']]

# --- 7d. Safety-Critical by Dataset – vs. Nurse Triage ---
ds_totals = df.groupby('dataset').size()
cross7_nt = sc_cases.groupby('dataset').size().reindex([1,2,3], fill_value=0).reset_index()
cross7_nt.columns = ['Dataset', 'Safety-Critical Count']
cross7_nt['Total in Dataset'] = cross7_nt['Dataset'].map(ds_totals)
cross7_nt['% of Dataset'] = (cross7_nt['Safety-Critical Count'] / cross7_nt['Total in Dataset'] * 100).round(1)
pd.DataFrame([['7d. Safety-Critical by Dataset – vs. Nurse Triage']]).to_excel(
    writer, sheet_name=sn7, startrow=r7, index=False, header=False); r7 += 1
cross7_nt.to_excel(writer, sheet_name=sn7, startrow=r7, index=False); r7 += len(cross7_nt)+2

# --- 7e. Safety-Critical by Dataset – vs. Clinician-Adjudicated ---
cross7_cc = sc_cc_cases.groupby('dataset').size().reindex([1,2,3], fill_value=0).reset_index()
cross7_cc.columns = ['Dataset', 'Safety-Critical Count']
cross7_cc['Total in Dataset'] = cross7_cc['Dataset'].map(ds_totals)
cross7_cc['% of Dataset'] = (cross7_cc['Safety-Critical Count'] / cross7_cc['Total in Dataset'] * 100).round(1)
pd.DataFrame([['7e. Safety-Critical by Dataset – vs. Clinician-Adjudicated']]).to_excel(
    writer, sheet_name=sn7, startrow=r7, index=False, header=False); r7 += 1
cross7_cc.to_excel(writer, sheet_name=sn7, startrow=r7, index=False); r7 += len(cross7_cc)+2

# --- 7f. Safety-Critical Under-Triage Rate by Condition & Reference (distance >= 2 AND under-triage) ---
pd.DataFrame([['7f. Safety-Critical Under-Triage Rate (distance >= 2 AND under-triage), % of all cases']]).to_excel(
    writer, sheet_name=sn7, startrow=r7, index=False, header=False); r7 += 1
def _sc_under(lo, hi, ref):
    d = np.array(row_dist(df[lo], df[hi], df[ref]), float)
    s = np.array(row_signed(df[lo], df[hi], df[ref]), float)
    m = ~np.isnan(d); d, s = d[m], s[m]; n = len(d)
    k = int(((d >= 2) & (s < 0)).sum())
    c = binomtest(k, n, 0.5).proportion_ci(0.95, method='exact')
    return [k, n, f"{k/n*100:.1f}%", f"({c.low*100:.1f}%, {c.high*100:.1f}%)"]
sc7f = []
for cn, lo, hi in [('Natural', 'nat_lo', 'nat_hi'), ('Multiturn', 'mt_lo', 'mt_hi')]:
    for rn, rc in [('Nurse Triage', 'gold_ord'), ('Clinician-Adjudicated', 'cc_cons_ord')]:
        sc7f.append([f'{cn} vs. {rn}'] + _sc_under(lo, hi, rc))
pd.DataFrame(sc7f, columns=['Condition / Reference', 'Safety-Critical Under-triage (n)',
                            'Total N', 'Rate (%)', '95% CI']).to_excel(
    writer, sheet_name=sn7, startrow=r7, index=False); r7 += len(sc7f)+2

style_sheet(writer.sheets[sn7], title_row=1, header_rows=[4])
print(f"  Step 7 complete. SC vs NT={len(sc_cases)}, SC vs ClinAdj={len(sc_cc_cases)}, severe (NT)={len(sev_either)}")

# =====================================================================
# STEP 8: PROMPT COUNT AS PREDICTOR
# =====================================================================
print("Running Step 8...")
sn8 = 'SECONDARY - Prompt Count'
r8 = 0
pd.DataFrame([['STEP 8: PROMPT COUNT AS PREDICTOR OF AGREEMENT']]).to_excel(
    writer, sheet_name=sn8, startrow=r8, index=False, header=False); r8 += 1
pd.DataFrame([[
    "Step 8 tests whether more conversational turns correlate with greater clinical distance. "
    "Spearman rank correlation (with Fisher z 95% CI) is reported, pooled and per dataset."
]]).to_excel(writer, sheet_name=sn8, startrow=r8, index=False, header=False); r8 += 2

def prompt_analysis(df_sub, count_col, dist_col, label):
    mask = (~df_sub[count_col].isna()) & (~df_sub[dist_col].isna())
    sub  = df_sub[mask]
    x, y = sub[count_col].values.astype(float), sub[dist_col].values.astype(float)
    n = len(x)
    rho, p = spearmanr(x, y)
    ci_lo, ci_hi = spearman_ci(rho, n)
    return [
        (f'{label} – N', n),
        (f'{label} – Spearman rho', f"{rho:.3f}"),
        (f'{label} – p-value', f"{p:.4f}"),
        (f'{label} – 95% CI', f"({ci_lo:.3f}, {ci_hi:.3f})" if not np.isnan(ci_lo) else 'N/A'),
    ]

rows8 = (prompt_analysis(df, 'nat_count', 'nat_dist', '8a. Natural Overall') +
         [('','')]*1 +
         prompt_analysis(df, 'mt_count',  'mt_dist',  '8b. Multiturn Overall'))
for ds in [1,2,3]:
    sub = df[df['dataset']==ds]
    rows8 += [('','')] + prompt_analysis(sub, 'nat_count', 'nat_dist', f'8e. Natural Dataset {ds}')
    rows8 += prompt_analysis(sub, 'mt_count',  'mt_dist',  f'8e. Multiturn Dataset {ds}')

pd.DataFrame(rows8, columns=['Metric','Value']).to_excel(writer, sheet_name=sn8, startrow=r8, index=False)
style_sheet(writer.sheets[sn8], title_row=1, header_rows=[4])
print("  Step 8 complete.")

# =====================================================================
# STEP 9: CHIEF COMPLAINT CATEGORY ANALYSIS
# =====================================================================
print("Running Step 9...")
sn9 = 'SECONDARY - By Complaint'
r9 = 0
pd.DataFrame([['STEP 9: CHIEF COMPLAINT CATEGORY ANALYSIS']]).to_excel(
    writer, sheet_name=sn9, startrow=r9, index=False, header=False); r9 += 1
pd.DataFrame([[
    "Step 9 groups complaints into 9 clinical categories. "
    "⚠ flags categories with <5 cases (underpowered). Ranked by Natural under-triage rate."
]]).to_excel(writer, sheet_name=sn9, startrow=r9, index=False, header=False); r9 += 2

lookup = df[['complaint','category']].drop_duplicates().sort_values('complaint')
pd.DataFrame([['Category Lookup Table']]).to_excel(writer, sheet_name=sn9, startrow=r9, index=False, header=False); r9 += 1
lookup.to_excel(writer, sheet_name=sn9, startrow=r9, index=False); r9 += len(lookup)+2

cat_rows = []
for cat in sorted(df['category'].unique()):
    sub = df[df['category']==cat]
    n   = len(sub)
    mn  = compute_metrics(sub['nat_lo'], sub['nat_hi'], sub['gold_ord'])
    mm  = compute_metrics(sub['mt_lo'],  sub['mt_hi'],  sub['gold_ord'])
    cat_rows.append({
        'Category': cat, 'N': n,
        'Nat Agree%':  f"{mn.get('agree_pct',0):.1f}%",
        'Nat MeanDist':f"{mn.get('mean_dist',0):.3f}",
        'Nat Under%':  f"{mn.get('under_pct',0):.1f}%",
        'MT Agree%':   f"{mm.get('agree_pct',0):.1f}%",
        'MT MeanDist': f"{mm.get('mean_dist',0):.3f}",
        'MT Under%':   f"{mm.get('under_pct',0):.1f}%",
        'Note': '⚠ Underpowered (<5)' if n < 5 else '',
        '_sort': mn.get('under_pct', 0),
    })
cat_df = pd.DataFrame(cat_rows).sort_values('_sort', ascending=False).drop('_sort', axis=1)
pd.DataFrame([['9b-d. Per-Category Metrics (ranked by Natural Under-triage Rate)']]).to_excel(
    writer, sheet_name=sn9, startrow=r9, index=False, header=False); r9 += 1
cat_df.to_excel(writer, sheet_name=sn9, startrow=r9, index=False)
style_sheet(writer.sheets[sn9], title_row=1, header_rows=[4])
print("  Step 9 complete.")

# =====================================================================
# STEP 10: GRADER RELIABILITY
# =====================================================================
print("Running Step 10...")
sn10 = 'SUPP - Grader Reliability'
r10 = 0
pd.DataFrame([['STEP 10: GRADER-LEVEL RELIABILITY ANALYSIS']]).to_excel(
    writer, sheet_name=sn10, startrow=r10, index=False, header=False); r10 += 1
pd.DataFrame([[
    "Step 10 reports per-grader metrics vs. both reference standards: Nurse Triage (NT) and Clinician-Adjudicated (ClinAdj). "
    "No prompt overlap between graders → pairwise inter-grader kappa and Fleiss' kappa cannot be computed. "
    "Graders with weighted kappa >1 SD below the group mean (for either reference) are flagged as potential outliers."
]]).to_excel(writer, sheet_name=sn10, startrow=r10, index=False, header=False); r10 += 2

grader_rows = []
kappas_nat, kappas_mt = [], []
kappas_nat_cc, kappas_mt_cc = [], []
for g in sorted(df['grader'].unique()):
    sub = df[df['grader']==g]
    mn   = compute_metrics(sub['nat_lo'], sub['nat_hi'], sub['gold_ord'])
    mm   = compute_metrics(sub['mt_lo'],  sub['mt_hi'],  sub['gold_ord'])
    mncc = compute_metrics(sub['nat_lo'], sub['nat_hi'], sub['cc_cons_ord'])
    mmcc = compute_metrics(sub['mt_lo'],  sub['mt_hi'],  sub['cc_cons_ord'])
    grader_rows.append({
        'Grader': g, 'N cases': mn.get('n',0),
        'Note': '(self-review — same person as reviewer)' if g==5 else '',
        'Nat Agree% (NT)':  f"{mn.get('agree_pct',0):.1f}%",
        'Nat MeanDist (NT)':f"{mn.get('mean_dist',0):.3f}",
        'Nat Under% (NT)':  f"{mn.get('under_pct',0):.1f}%",
        'Nat Kappa (NT)':   f"{mn.get('kappa',np.nan):.3f}" if not np.isnan(mn.get('kappa',np.nan)) else 'N/A',
        'MT Agree% (NT)':   f"{mm.get('agree_pct',0):.1f}%",
        'MT MeanDist (NT)': f"{mm.get('mean_dist',0):.3f}",
        'MT Under% (NT)':   f"{mm.get('under_pct',0):.1f}%",
        'MT Kappa (NT)':    f"{mm.get('kappa',np.nan):.3f}" if not np.isnan(mm.get('kappa',np.nan)) else 'N/A',
        'Nat Agree% (ClinAdj)':  f"{mncc.get('agree_pct',0):.1f}%",
        'Nat MeanDist (ClinAdj)':f"{mncc.get('mean_dist',0):.3f}",
        'Nat Under% (ClinAdj)':  f"{mncc.get('under_pct',0):.1f}%",
        'Nat Kappa (ClinAdj)':   f"{mncc.get('kappa',np.nan):.3f}" if not np.isnan(mncc.get('kappa',np.nan)) else 'N/A',
        'MT Agree% (ClinAdj)':   f"{mmcc.get('agree_pct',0):.1f}%",
        'MT MeanDist (ClinAdj)': f"{mmcc.get('mean_dist',0):.3f}",
        'MT Under% (ClinAdj)':   f"{mmcc.get('under_pct',0):.1f}%",
        'MT Kappa (ClinAdj)':    f"{mmcc.get('kappa',np.nan):.3f}" if not np.isnan(mmcc.get('kappa',np.nan)) else 'N/A',
        '_nk': mn.get('kappa', np.nan),   '_mk': mm.get('kappa', np.nan),
        '_nkcc': mncc.get('kappa', np.nan), '_mkcc': mmcc.get('kappa', np.nan),
    })
    if not np.isnan(mn.get('kappa', np.nan)):   kappas_nat.append(mn['kappa'])
    if not np.isnan(mm.get('kappa', np.nan)):   kappas_mt.append(mm['kappa'])
    if not np.isnan(mncc.get('kappa', np.nan)): kappas_nat_cc.append(mncc['kappa'])
    if not np.isnan(mmcc.get('kappa', np.nan)): kappas_mt_cc.append(mmcc['kappa'])

mean_nat_k = np.nanmean(kappas_nat); sd_nat_k = np.nanstd(kappas_nat)
mean_mt_k  = np.nanmean(kappas_mt);  sd_mt_k  = np.nanstd(kappas_mt)
mean_nat_kcc = np.nanmean(kappas_nat_cc); sd_nat_kcc = np.nanstd(kappas_nat_cc)
mean_mt_kcc  = np.nanmean(kappas_mt_cc);  sd_mt_kcc  = np.nanstd(kappas_mt_cc)
for row in grader_rows:
    flags = []
    if not np.isnan(row['_nk'])   and row['_nk']   < mean_nat_k   - sd_nat_k:   flags.append('Nat κ outlier (NT)')
    if not np.isnan(row['_mk'])   and row['_mk']   < mean_mt_k    - sd_mt_k:    flags.append('MT κ outlier (NT)')
    if not np.isnan(row['_nkcc']) and row['_nkcc'] < mean_nat_kcc - sd_nat_kcc: flags.append('Nat κ outlier (ClinAdj)')
    if not np.isnan(row['_mkcc']) and row['_mkcc'] < mean_mt_kcc  - sd_mt_kcc:  flags.append('MT κ outlier (ClinAdj)')
    row['Outlier Flag'] = '; '.join(flags) if flags else ''
    for k in ('_nk','_mk','_nkcc','_mkcc'): del row[k]

pd.DataFrame(grader_rows).to_excel(writer, sheet_name=sn10, startrow=r10, index=False); r10 += len(grader_rows)+2
pd.DataFrame([[
    f'Group means – vs Nurse Triage: Nat κ {mean_nat_k:.3f} (SD={sd_nat_k:.3f}), MT κ {mean_mt_k:.3f} (SD={sd_mt_k:.3f}) | '
    f'vs Clinician-Adjudicated: Nat κ {mean_nat_kcc:.3f} (SD={sd_nat_kcc:.3f}), MT κ {mean_mt_kcc:.3f} (SD={sd_mt_kcc:.3f})'
]]).to_excel(writer, sheet_name=sn10, startrow=r10, index=False, header=False); r10 += 3

# --- 10b. Sensitivity: primary endpoints excluding flagged grader (Grader 3) ---
pd.DataFrame([['10b. Sensitivity Analysis – Primary Endpoints Excluding Flagged Grader (Grader 3)']]).to_excel(
    writer, sheet_name=sn10, startrow=r10, index=False, header=False); r10 += 1
pd.DataFrame([[
    "Grader 3 (n=43) fell >1 SD below the group-mean weighted kappa in both conditions and was flagged "
    "as a potential outlier. Primary endpoints are recomputed excluding Grader 3 (N=212) and compared "
    "with the full cohort (N=255). Under-triage % and its binomial p-value are computed among disagreement cases."
]]).to_excel(writer, sheet_name=sn10, startrow=r10, index=False, header=False); r10 += 2

_df_excl = df[df['grader'].astype(str).str.strip() != '3']
def _sens_ep(d, lo, hi, ref):
    dist = np.array(row_dist(d[lo], d[hi], d[ref]), float)
    s    = np.array(row_signed(d[lo], d[hi], d[ref]), float)
    m = ~np.isnan(dist); dist, s = dist[m], s[m]; n = len(dist)
    nerr = int((s != 0).sum()); nu = int((s < 0).sum())
    agree = (dist == 0).mean()*100
    pu = nu/nerr*100 if nerr else np.nan
    pv = binomtest(nu, nerr, 0.5).pvalue if nerr else np.nan
    pdisp = '<0.0001' if (nerr and pv < 0.0001) else (f"{pv:.4f}" if nerr else 'N/A')
    return n, agree, dist.mean(), nu, nerr, pu, pdisp

sens_rows = []
for cn, lo, hi in [('Natural','nat_lo','nat_hi'), ('Multiturn','mt_lo','mt_hi')]:
    for rn, rc in [('Nurse Triage','gold_ord'), ('Clinician-Adjudicated','cc_cons_ord')]:
        nf, agf, mdf, nuf, nef, puf, pvf = _sens_ep(df,       lo, hi, rc)
        ne, age, mde, nue, nee, pue, pve = _sens_ep(_df_excl, lo, hi, rc)
        sens_rows.append([
            f'{cn} vs. {rn}',
            f"{agf:.1f}%", f"{age:.1f}%",
            f"{mdf:.2f}",  f"{mde:.2f}",
            f"{puf:.1f}% ({nuf}/{nef})", pvf,
            f"{pue:.1f}% ({nue}/{nee})", pve])
sens_cols = ['Comparison', 'Agree% (full N=255)', 'Agree% (excl G3 N=212)',
             'Mean Dist (full)', 'Mean Dist (excl G3)',
             'Under-triage% (full)', 'Under p (full)',
             'Under-triage% (excl G3)', 'Under p (excl G3)']
pd.DataFrame(sens_rows, columns=sens_cols).to_excel(writer, sheet_name=sn10, startrow=r10, index=False)
r10 += len(sens_rows) + 2

style_sheet(writer.sheets[sn10], title_row=1, header_rows=[4])
print("  Step 10 complete.")

# =====================================================================
# STEP 11: DIRECTION-OF-ERROR ASYMMETRY
# =====================================================================
print("Running Step 11...")
sn11 = 'PRIMARY - Error Direction'
r11 = 0
pd.DataFrame([['STEP 11: DIRECTION-OF-ERROR ASYMMETRY TEST']]).to_excel(
    writer, sheet_name=sn11, startrow=r11, index=False, header=False); r11 += 1
pd.DataFrame([[
    "Step 11 tests whether the LLM conditions systematically err toward lower-acuity care (under-triage). "
    "Among disagreement cases ('errors'), a two-sided binomial test evaluates whether the proportion that "
    "are under-triage differs from 0.5; the 95% CI is the exact (Clopper–Pearson) interval on P(under | error). "
    "Under-triage counts only when the entire decision range is below the reference level. "
    "Section 11A uses Nurse Triage as the reference standard; Section 11B uses Clinician-Adjudicated. "
    "Restricted rows limit to reference acuity C or D, where under-triage is most safety-relevant."
]]).to_excel(writer, sheet_name=sn11, startrow=r11, index=False, header=False); r11 += 2

def asymmetry_row(signed_col, ref_col, label, df_sub=None):
    if df_sub is None: df_sub = df
    mask = (~df_sub[signed_col].isna()) & (~df_sub[ref_col].isna())
    s = df_sub[mask][signed_col].values
    n_err   = int((s != 0).sum())
    n_under = int((s  < 0).sum())
    n_over  = int((s  > 0).sum())
    if n_err == 0:
        return [label, 0, 0, 0, 'N/A', 'N/A', 'N/A', 'No']
    p_under = n_under / n_err
    res = binomtest(n_under, n_err, p=0.5, alternative='two-sided')
    ci  = res.proportion_ci(confidence_level=0.95)
    rej = 'Yes' if res.pvalue < 0.05 else 'No'
    p_disp = "<0.0001" if res.pvalue < 0.0001 else f"{res.pvalue:.4f}"
    return [label, n_err, n_under, n_over, f"{p_under*100:.1f}%",
            f"({ci.low*100:.1f}%, {ci.high*100:.1f}%)", p_disp, rej]

asym_cols = ['Condition','N Errors','N Under','N Over','% Under-triage',
             '95% CI (P under|error)','Binomial p-value','Reject H0 (p<0.05)?']

# --- 11A: vs. Nurse Triage ---
pd.DataFrame([['11A: Under-triage Direction Bias vs. Nurse Triage']]).to_excel(
    writer, sheet_name=sn11, startrow=r11, index=False, header=False); r11 += 1
hdr_a = r11 + 1
rows_a = []
for restrict in [False, True]:
    sub_cd  = df[df['gold_ord'].isin([2.0, 3.0])] if restrict else df
    lbl_sfx = ' (Nurse Triage = C/D only)' if restrict else ''
    rows_a.append(asymmetry_row('nat_signed', 'gold_ord', f'Natural{lbl_sfx}', sub_cd))
    rows_a.append(asymmetry_row('mt_signed',  'gold_ord', f'Multiturn{lbl_sfx}', sub_cd))
    rows_a.append(asymmetry_row('cc_signed',  'gold_ord', f'Clinician-Adjudicated (benchmark){lbl_sfx}', sub_cd))
pd.DataFrame(rows_a, columns=asym_cols).to_excel(writer, sheet_name=sn11, startrow=r11, index=False)
r11 += len(rows_a) + 2

# --- 11B: vs. Clinician-Adjudicated ---
pd.DataFrame([['11B: Under-triage Direction Bias vs. Clinician-Adjudicated']]).to_excel(
    writer, sheet_name=sn11, startrow=r11, index=False, header=False); r11 += 1
hdr_b = r11 + 1
rows_b = []
for restrict in [False, True]:
    sub_cc  = df[df['cc_cons_ord'].isin([2.0, 3.0])] if restrict else df
    lbl_sfx = ' (Clinician-Adjudicated = C/D only)' if restrict else ''
    rows_b.append(asymmetry_row('nat_signed_cc', 'cc_cons_ord', f'Natural{lbl_sfx}', sub_cc))
    rows_b.append(asymmetry_row('mt_signed_cc',  'cc_cons_ord', f'Multiturn{lbl_sfx}', sub_cc))
pd.DataFrame(rows_b, columns=asym_cols).to_excel(writer, sheet_name=sn11, startrow=r11, index=False)
r11 += len(rows_b) + 2

style_sheet(writer.sheets[sn11], title_row=1, header_rows=[hdr_a, hdr_b])
print(f"  Step 11 complete. Nat/NT under p={asymmetry_row('nat_signed','gold_ord','x')[6]}, "
      f"Nat/ClinAdj under p={asymmetry_row('nat_signed_cc','cc_cons_ord','x')[6]}")

# =====================================================================
# STEP 12: MASTER SUMMARY
# =====================================================================
print("Running Step 12...")
sn12 = 'SUPP - Summary Table'
r12 = 0
pd.DataFrame([['STEP 12: MASTER SUMMARY TABLE']]).to_excel(
    writer, sheet_name=sn12, startrow=r12, index=False, header=False); r12 += 1
pd.DataFrame([[
    "Step 12 consolidates key metrics. *** flags where BOTH LLM conditions perform worse than Clinician-Adjudicated."
]]).to_excel(writer, sheet_name=sn12, startrow=r12, index=False, header=False); r12 += 2

nat_a_v = (df['nat_dist'] == 0).values
mt_a_v  = (df['mt_dist']  == 0).values
_, mcn_p6 = mcnemar_test(nat_a_v, mt_a_v)

def binom_p_from_signed(signed_arr):
    s = np.array(signed_arr, float)
    n_err   = int((s != 0).sum())
    n_under = int((s  < 0).sum())
    if n_err == 0: return np.nan
    return binomtest(n_under, n_err, p=0.5, alternative='two-sided').pvalue

binom_nat = binom_p_from_signed(df['nat_signed'])
binom_mt  = binom_p_from_signed(df['mt_signed'])
binom_cc  = binom_p_from_signed(df['cc_signed'])

def sc_rate(signed_arr, dist_arr, n_total):
    s = np.array(signed_arr, float); d = np.array(dist_arr, float)
    return ((s < 0) & (d >= 2)).sum() / n_total * 100

nat_sc = sc_rate(df['nat_signed'], df['nat_dist'], len(df))
mt_sc  = sc_rate(df['mt_signed'],  df['mt_dist'],  len(df))
cc_sc  = sc_rate(df['cc_signed'],  df['cc_dist'],  len(df))

rows12 = [
    ('N cases', m2a['n'], m2b['n']),
    ('Exact Agreement with Nurse Triage (%)',
     f"{m2a['agree_pct']:.1f}% ({m2a['agree_ci_lo']:.1f}–{m2a['agree_ci_hi']:.1f}%)",
     f"{m2b['agree_pct']:.1f}% ({m2b['agree_ci_lo']:.1f}–{m2b['agree_ci_hi']:.1f}%)"),
    ('Exact Agreement with Clinician-Adjudicated (%)',
     f"{m3b_nat['agree_pct']:.1f}% ({m3b_nat['agree_ci_lo']:.1f}–{m3b_nat['agree_ci_hi']:.1f}%)",
     f"{m3b_mt['agree_pct']:.1f}% ({m3b_mt['agree_ci_lo']:.1f}–{m3b_mt['agree_ci_hi']:.1f}%)"),
    ("Cohen's Weighted Kappa vs. Nurse Triage (95% CI)",
     f"{m2a['kappa']:.3f} ({m2a['kappa_lo']:.3f}, {m2a['kappa_hi']:.3f})",
     f"{m2b['kappa']:.3f} ({m2b['kappa_lo']:.3f}, {m2b['kappa_hi']:.3f})"),
    ("Cohen's Weighted Kappa vs. Clinician-Adjudicated (95% CI)",
     f"{m3b_nat['kappa']:.3f} ({m3b_nat['kappa_lo']:.3f}, {m3b_nat['kappa_hi']:.3f})",
     f"{m3b_mt['kappa']:.3f} ({m3b_mt['kappa_lo']:.3f}, {m3b_mt['kappa_hi']:.3f})"),
    ('Mean Distance from Nurse Triage (SD)',
     f"{m2a['mean_dist']:.3f} ({m2a['sd_dist']:.3f})",
     f"{m2b['mean_dist']:.3f} ({m2b['sd_dist']:.3f})"),
    ('Mean Distance from Clinician-Adjudicated (SD)',
     f"{m3b_nat['mean_dist']:.3f} ({m3b_nat['sd_dist']:.3f})",
     f"{m3b_mt['mean_dist']:.3f} ({m3b_mt['sd_dist']:.3f})"),
    ('Under-triage Rate vs. Nurse Triage (%)',
     f"{m2a['under_pct']:.1f}%", f"{m2b['under_pct']:.1f}%"),
    ('Under-triage Rate vs. Clinician-Adjudicated (%)',
     f"{m3b_nat['under_pct']:.1f}%", f"{m3b_mt['under_pct']:.1f}%"),
    ('Under-triage Safety-Critical vs. Nurse Triage (Dist≥2) (%)',
     f"{nat_sc:.1f}%", f"{mt_sc:.1f}%"),
    ('Over-triage Rate vs. Nurse Triage (%)',
     f"{m2a['over_pct']:.1f}%", f"{m2b['over_pct']:.1f}%"),
    ('Binomial p-value – Direction Asymmetry vs. Nurse Triage',
     f"{binom_nat:.4f}", f"{binom_mt:.4f}"),
    ('Natural vs. Multiturn McNemar p-value',
     f"{mcn_p6:.4f}" if not np.isnan(mcn_p6) else 'N/A', '(same test)'),
]
pd.DataFrame(rows12, columns=['Metric','Natural','Multiturn']).to_excel(
    writer, sheet_name=sn12, startrow=r12, index=False); r12 += len(rows12)+2

for note in [
    f"Kappa vs. Nurse Triage: Natural={kappa_label(m2a['kappa'])} | Multiturn={kappa_label(m2b['kappa'])}",
    'Kappa benchmarks: <0.20=Slight, 0.20-0.40=Fair, 0.40-0.60=Moderate, 0.60-0.80=Substantial, 0.80+=Almost Perfect',
    'Distance benchmarks: 0.00-0.25=Excellent, 0.26-0.50=Good, 0.51-1.00=Moderate, >1.00=Poor (patient safety concern)',
]:
    pd.DataFrame([[note]]).to_excel(writer, sheet_name=sn12, startrow=r12, index=False, header=False); r12 += 1
style_sheet(writer.sheets[sn12], title_row=1, header_rows=[4])

# =====================================================================
# COMPARISON SUMMARY TABLE
# =====================================================================
# Natural vs. Multiturn (against each other — Multiturn as reference)
nat_signed_vs_mt = row_signed(df['nat_lo'], df['nat_hi'], df['mt_lo'])

valid_nat_mt = (~df['nat_lo'].isna()) & (~df['mt_lo'].isna())
valid_nat_cc = (~df['nat_lo'].isna()) & (~df['cc_cons_ord'].isna())
valid_mt_cc  = (~df['mt_lo'].isna())  & (~df['cc_cons_ord'].isna())
valid_nat_nu = (~df['nat_lo'].isna()) & (~df['gold_ord'].isna())
valid_mt_nu  = (~df['mt_lo'].isna())  & (~df['gold_ord'].isna())

def comp_summary_row(label, decision_label, reference_label, signed_arr, valid_mask):
    s = np.array(signed_arr, float)[valid_mask]
    n = int(valid_mask.sum())
    return [label, decision_label, reference_label, n,
            f"{(s == 0).mean()*100:.1f}%",
            f"{(s < 0).mean()*100:.1f}%",
            f"{(s > 0).mean()*100:.1f}%"]

comp_rows = [
    comp_summary_row('Natural vs. Multiturn',
                     'GPT Natural', 'GPT Multiturn',
                     nat_signed_vs_mt, valid_nat_mt),
    comp_summary_row('Natural vs. Clinician-Adjudicated',
                     'GPT Natural', 'Clinician-Adjudicated',
                     nat_signed_cc, valid_nat_cc),
    comp_summary_row('Multiturn vs. Clinician-Adjudicated',
                     'GPT Multiturn', 'Clinician-Adjudicated',
                     mt_signed_cc, valid_mt_cc),
    comp_summary_row('Natural vs. Nurse Triage',
                     'GPT Natural', 'Nurse Triage',
                     df['nat_signed'], valid_nat_nu),
    comp_summary_row('Multiturn vs. Nurse Triage',
                     'GPT Multiturn', 'Nurse Triage',
                     df['mt_signed'], valid_mt_nu),
]
comp_cols = ['Comparison', 'Decision', 'Reference', 'N',
             'Exact Agreement', 'Under-triage', 'Over-triage']
df_comp = pd.DataFrame(comp_rows, columns=comp_cols)

sn_comp = 'SUPP - Comparison Summary'
ws_comp = writer.book.create_sheet(sn_comp)
writer.sheets[sn_comp] = ws_comp

ws_comp['A1'] = 'COMPARISON SUMMARY: Exact Agreement, Under-triage, and Over-triage'
ws_comp.merge_cells('A1:G1')

for c_idx, col_name in enumerate(comp_cols, start=1):
    ws_comp.cell(row=2, column=c_idx, value=col_name)

for r_idx, row_vals in enumerate(comp_rows, start=3):
    for c_idx, val in enumerate(row_vals, start=1):
        ws_comp.cell(row=r_idx, column=c_idx, value=val)

# Footnote row
fn_row = 3 + len(comp_rows) + 1
ws_comp.cell(row=fn_row, column=1, value=(
    "Notes: Under-triage = decision urgency < reference urgency.  "
    "Over-triage = decision urgency > reference urgency.  "
    "For Natural vs. Multiturn, Multiturn is the reference (under-triage = Natural assigned lower urgency than Multiturn).  "
    "Range values use oracle assignment (clamp to nearest reference value).  "
    "Clinician-Adjudicated uses conservative (upper urgency) bound for ambiguous values."
))
ws_comp.merge_cells(start_row=fn_row, start_column=1, end_row=fn_row, end_column=7)
fn_cell = ws_comp.cell(row=fn_row, column=1)
fn_cell.font = Font(name=SHEET_FONT, size=9, italic=True)
fn_cell.alignment = Alignment(wrap_text=True, vertical='top')
ws_comp.row_dimensions[fn_row].height = 50

# Styling (booktabs)
style_sheet(ws_comp, title_row=1, header_rows=[2])
col_widths = {'A': 38, 'B': 22, 'C': 32, 'D': 6, 'E': 18, 'F': 16, 'G': 15}
for col_letter, w in col_widths.items():
    ws_comp.column_dimensions[col_letter].width = w
ws_comp.row_dimensions[1].height = 22
ws_comp.row_dimensions[2].height = 28

for row in ws_comp.iter_rows(min_row=3, max_row=3+len(comp_rows)-1):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    row[0].alignment = Alignment(horizontal='left', vertical='center')
    row[1].alignment = Alignment(horizontal='left', vertical='center')
    row[2].alignment = Alignment(horizontal='left', vertical='center')

print("  Comparison Summary table written.")

# Reorder tabs to reflect primary vs. secondary outcomes: PRIMARY → SECONDARY → SUPP
_desired_order = [
    'Table 1',
    'PRIMARY - Performance',
    'PRIMARY - Safety-Critical',
    'PRIMARY - Error Direction',
    'SECONDARY - Sensitivity',
    'SECONDARY - By Dataset',
    'SECONDARY - Dataset x Level',
    'SECONDARY - By Complaint',
    'SECONDARY - Prompt Count',
    'SECONDARY - Nat vs Multiturn',
    'SUPP - Descriptives',
    'SUPP - ClinAdj vs Nurse',
    'SUPP - Grader Reliability',
    'SUPP - Summary Table',
    'SUPP - Comparison Summary',
]
_wb = writer.book
_present = {ws.title: ws for ws in _wb.worksheets}
_wb._sheets = ([_present[name] for name in _desired_order if name in _present] +
               [ws for ws in _wb.worksheets if ws.title not in _desired_order])
_wb.active = 0

print(f"\nWriting to {OUTPUT_FILE}...")
writer.close()
print(f"Done. Output: {OUTPUT_FILE}")
